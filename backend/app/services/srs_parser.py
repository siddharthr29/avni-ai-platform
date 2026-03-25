"""Parse Avni SRS Excel files into structured SRSData for bundle generation.

Supports two known SRS Excel layouts:
  - Simple (Sangwari-style): Group/Page Name, Field Name, Data Type, Mandatory,
    When to show, When NOT to show, Unique option, Default Value, OPTIONS, Validation
  - Expanded (Gubbachi-style): Page Name, Field Name, Data Type, Mandatory,
    User/System, Negative, Decimal, Min/Max, Unit, Current Date, Future Date,
    Past Date, Selection Type, OPTIONS, Unique option, Validation, Show, Not Show

Both layouts are auto-detected by inspecting the header row of each sheet.
"""

from __future__ import annotations

import logging
import re
from pathlib import Path
from typing import Any

import openpyxl
from openpyxl.worksheet.worksheet import Worksheet

from app.models.schemas import (
    SRSData,
    SRSFormDefinition,
    SRSFormField,
    SRSFormGroup,
)

logger = logging.getLogger(__name__)

# ---------------------------------------------------------------------------
# Constants
# ---------------------------------------------------------------------------

# Mapping from SRS data-type strings (case-insensitive) to Avni concept types
_DATA_TYPE_MAP: dict[str, str] = {
    # Text variants
    "text": "Text",
    "short text": "Text",
    "long text": "Notes",
    "notes": "Notes",
    "note": "Notes",
    # Numeric
    "number": "Numeric",
    "numeric": "Numeric",
    "integer": "Numeric",
    "decimal": "Numeric",
    # Date / Time
    "date": "Date",
    "calender": "Date",
    "calendar": "Date",
    "datetime": "DateTime",
    "time": "Time",
    # Coded (single/multi select)
    "single select": "Coded",
    "multi select": "Coded",
    "multiselect": "Coded",
    "singleselect": "Coded",
    "dropdown": "Coded",
    "pre added options": "Coded",
    "pre-added options": "Coded",
    "coded": "Coded",
    # Media
    "image": "Image",
    "file": "File",
    "video": "Video",
    "audio": "Audio",
    # Phone
    "phone number": "PhoneNumber",
    "phonenumber": "PhoneNumber",
    "phone": "PhoneNumber",
    # Location
    "location": "Location",
    # Id
    "id": "Id",
    # Subject
    "subject": "Subject",
    # Question group
    "question group": "QuestionGroup",
    # Special
    "auto calculated": "Numeric",
}

# Known non-form sheet name patterns (case-insensitive substring matching)
_NON_FORM_PATTERNS: list[str] = [
    "help",
    "status tracker",
    "summary",
    "overview",
    "user persona",
    "user type",
    "w3h",
    "report",
    "bi report",
    "dashboard",
    "card",
    "permission",
    "privilege",
    "discussion",
    "decision",
    "issue",
    "review",
    "checklist",
    "modelling",
    "modeling",
    "location hierarchy",
    "subject type",
    "program encounter",
    "encounter",
    "other important",
]


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------


def _normalize(val: Any) -> str | None:
    """Return stripped string or None."""
    if val is None:
        return None
    s = str(val).strip()
    return s if s else None


def _normalize_lower(val: Any) -> str:
    """Return lower-cased stripped string."""
    if val is None:
        return ""
    return str(val).strip().lower()


def _is_yes(val: Any) -> bool:
    """Check if a cell value means 'yes' / mandatory."""
    if val is None:
        return False
    s = str(val).strip().lower()
    return s in ("yes", "y", "true", "1", "mandatory")


def _parse_options(raw: Any) -> list[str]:
    """Extract option strings from a cell value.

    Handles multiple real-world formats:
      - "a). Male\\nb). Female"
      - "Active\\nClosed"
      - "Option1; Option2; Option3"
      - "Option1, Option2, Option3"
      - "Option1/Option2/Option3"  (only when clearly a list)
      - "Yes/No"
    """
    if raw is None:
        return []
    text = str(raw).strip()
    if not text:
        return []

    options: list[str] = []

    # Try inline letter-prefixed options: "a). Yes b). No" or "a) X b) Y"
    # This pattern appears when all options are on a single line
    if re.search(r"[a-z]\)\.", text, re.IGNORECASE):
        parts = re.split(r"\s*[a-z]\)\.\s*", text, flags=re.IGNORECASE)
        cleaned_parts = [p.strip().rstrip(",;") for p in parts if p.strip()]
        if len(cleaned_parts) >= 2:
            return cleaned_parts
    elif re.search(r"[a-z]\)", text, re.IGNORECASE) and "\n" not in text:
        parts = re.split(r"\s*[a-z]\)\s*", text, flags=re.IGNORECASE)
        cleaned_parts = [p.strip().rstrip(",;") for p in parts if p.strip()]
        if len(cleaned_parts) >= 2:
            return cleaned_parts

    # Try newline-separated (most common in real SRS files)
    lines = text.split("\n")
    if len(lines) > 1:
        for line in lines:
            line = line.strip()
            if not line:
                continue
            # Strip "a). ", "b). ", "1. ", "- " prefixes
            cleaned = re.sub(r"^[a-z0-9]+\)\.\s*", "", line, flags=re.IGNORECASE)
            cleaned = re.sub(r"^[a-z0-9]+\)\s*", "", cleaned, flags=re.IGNORECASE)
            cleaned = re.sub(r"^[-\u2022*]\s*", "", cleaned)
            cleaned = re.sub(r"^\d+\.\s*", "", cleaned)
            cleaned = cleaned.strip().rstrip(",").strip()
            if cleaned:
                options.append(cleaned)
        if options:
            return options

    # Try semicolon-separated
    if ";" in text:
        for part in text.split(";"):
            part = part.strip().rstrip(",").strip()
            if part:
                options.append(part)
        if options:
            return options

    # Try slash-separated for short lists like "Yes/No", "Alcohol/Tobacco/None"
    if "/" in text and len(text) < 200:
        parts = [p.strip() for p in text.split("/") if p.strip()]
        if all(len(p) < 50 for p in parts) and len(parts) >= 2:
            return parts

    # Try comma-separated (last resort, risky for text with commas)
    if "," in text:
        parts = [p.strip().rstrip(",").strip() for p in text.split(",") if p.strip()]
        if len(parts) >= 2:
            return parts

    # Single value - strip any prefix
    cleaned = re.sub(r"^[a-z0-9]+\)\.\s*", "", text, flags=re.IGNORECASE)
    cleaned = cleaned.strip()
    if cleaned:
        return [cleaned]

    return []


def _parse_min_max(raw: Any) -> tuple[float | None, float | None]:
    """Try to extract min and max numeric bounds from a validation cell."""
    if raw is None:
        return None, None
    text = str(raw).strip()

    # Pattern: "Min: 0, Max: 200" or "0-200" or "Min 0 Max 200"
    m = re.search(r"min\s*[:=]?\s*(-?[\d.]+)", text, re.IGNORECASE)
    low = float(m.group(1)) if m else None

    m = re.search(r"max\s*[:=]?\s*(-?[\d.]+)", text, re.IGNORECASE)
    high = float(m.group(1)) if m else None

    if low is None and high is None:
        # Try "0-200" range pattern
        m = re.match(r"^(-?[\d.]+)\s*[-\u2013]\s*(-?[\d.]+)$", text)
        if m:
            low = float(m.group(1))
            high = float(m.group(2))

    return low, high


def _parse_unit(raw: Any) -> str | None:
    """Extract a unit string from a cell."""
    if raw is None:
        return None
    text = str(raw).strip()
    if not text:
        return None
    # Sometimes unit is in parentheses within the field name like "Weight (kg)"
    # This function handles explicit unit cells
    return text if text else None


def _extract_unit_from_name(name: str) -> tuple[str, str | None]:
    """Extract unit from field names like 'Weight (kg)' or 'MUAC (cm)'.

    Returns (cleaned_name, unit_or_none).
    """
    m = re.search(r"\(([a-zA-Z%/]+)\)\s*$", name)
    if m:
        unit = m.group(1).strip()
        cleaned = name[: m.start()].strip()
        return cleaned, unit
    return name, None


def _is_form_header_row(row: tuple) -> bool:
    """Check if a row looks like a form field header row."""
    # Normalize first few non-None values
    texts = []
    for val in row:
        if val is not None:
            texts.append(_normalize_lower(val))
        if len(texts) >= 5:
            break

    needed = {"field name", "data type"}
    found = set()
    for t in texts:
        for n in needed:
            if n in t:
                found.add(n)
    return len(found) >= 2


def _detect_column_indices(header_row: tuple) -> dict[str, int]:
    """Map logical column names to indices from the header row."""
    mapping: dict[str, int] = {}
    for i, val in enumerate(header_row):
        lower = _normalize_lower(val)
        if not lower:
            continue

        if "page" in lower or "group" in lower:
            mapping.setdefault("page_name", i)
        elif lower == "field name" or (lower.startswith("field") and "name" in lower):
            mapping.setdefault("field_name", i)
        elif lower == "data type" or lower.startswith("data type"):
            mapping.setdefault("data_type", i)
        elif "mandatory" in lower:
            mapping.setdefault("mandatory", i)
        elif "user enter" in lower or "system generat" in lower:
            mapping.setdefault("user_system", i)
        elif "negative" in lower:
            mapping.setdefault("allow_negative", i)
        elif "decimal" in lower:
            mapping.setdefault("allow_decimal", i)
        elif "max" in lower and "min" in lower and "limit" in lower:
            mapping.setdefault("min_max", i)
        elif "unit" in lower and "unique" not in lower:
            mapping.setdefault("unit", i)
        elif "selection type" in lower or "pre added options selection" in lower:
            mapping.setdefault("selection_type", i)
        elif "unique option" in lower:
            mapping.setdefault("unique_option", i)
        elif "option" in lower and ("needed" in lower or lower.startswith("options")):
            mapping.setdefault("options", i)
        elif lower.startswith("option") and "condition" not in lower and "unique" not in lower:
            mapping.setdefault("options", i)
        elif "when to show" in lower or (lower.startswith("when") and "not" not in lower):
            mapping.setdefault("show_when", i)
        elif "when not" in lower or "not to show" in lower:
            mapping.setdefault("hide_when", i)
        elif "default" in lower:
            mapping.setdefault("default_value", i)
        elif "validation" in lower or "condition" in lower:
            mapping.setdefault("validation", i)

    return mapping


def _classify_sheet(name: str) -> str:
    """Classify a sheet by its name. Returns one of:
    'summary', 'user', 'w3h', 'form', 'report', 'dashboard',
    'permission', 'modelling', 'skip', or 'unknown'.
    """
    lower = name.strip().lower()

    if "help" in lower or "status tracker" in lower:
        return "skip"
    if "form summary" in lower or "form overview" in lower:
        return "form_summary"
    if "summary" in lower or "overview" in lower:
        return "summary"
    if "user" in lower and ("type" in lower or "persona" in lower):
        return "user"
    if "w3h" in lower:
        return "w3h"
    if "report" in lower or "bi report" in lower:
        return "report"
    if "dashboard" in lower or "card" in lower:
        return "dashboard"
    if "permission" in lower or "privilege" in lower:
        return "permission"
    if "model" in lower:
        return "modelling"
    if "location" in lower and "hierarch" in lower:
        return "location"
    if "program encounter" in lower:
        return "program_encounters"
    if lower.strip() in ("program", "programs"):
        return "programs_meta"
    if lower.strip() in ("encounter", "encounters"):
        return "encounters_meta"
    if "subject type" in lower:
        return "subject_types"
    if "discussion" in lower or "decision" in lower or "issue" in lower:
        return "skip"
    if "review" in lower or "checklist" in lower:
        return "skip"
    if "other important" in lower:
        return "skip"
    if "visit" in lower and "schedul" in lower:
        return "visit_scheduling"
    if "app dashboard" in lower:
        return "dashboard"

    return "unknown"


# ---------------------------------------------------------------------------
# Sheet Parsers
# ---------------------------------------------------------------------------


def _parse_summary_sheet(ws: Worksheet) -> dict[str, Any]:
    """Parse a Project Summary / Program Summary sheet.

    Returns dict with keys: org_name, location_hierarchy, programs_text, etc.
    """
    result: dict[str, Any] = {}

    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, values_only=True):
        label = _normalize_lower(row[0]) if row and len(row) > 0 else ""
        value = _normalize(row[1]) if row and len(row) > 1 else None

        if not label or not value:
            continue

        if "organisation" in label or "organization" in label:
            result["org_name"] = value
        elif "program name" in label or "name of the program" in label:
            result["org_name"] = result.get("org_name") or value
            result["program_name"] = value
        elif "location" in label and ("hierarchy" in label or "heirarchy" in label):
            result["location_hierarchy"] = value
        elif "geographical" in label:
            result["geography"] = value
        elif "project" in label and "portfolio" in label:
            result["programs_text"] = value
        elif "number of user" in label:
            result["users_text"] = value
        elif "beneficiar" in label:
            result["beneficiaries_text"] = value
        elif "objective" in label:
            result["objective"] = value
        elif "name of the program" in label:
            result["program_name"] = value

    return result


def _parse_user_sheet(ws: Worksheet) -> list[str]:
    """Parse User Types / User Persona sheet. Returns list of group names."""
    groups: list[str] = []
    first_data_row = True

    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, values_only=True):
        user_type = _normalize(row[0]) if row and len(row) > 0 else None
        if not user_type:
            continue

        lower = user_type.lower()
        # Skip header row
        if "user type" in lower or "user persona" in lower:
            first_data_row = True
            continue

        # Extract just the primary name (before newlines/parentheses)
        name = user_type.split("\n")[0].strip()
        name = re.sub(r"\s*\(.*?\)\s*", "", name).strip()
        if name and name.lower() not in ("user type", "description", "number"):
            groups.append(name)

    return groups


def _parse_w3h_sheet(ws: Worksheet) -> list[dict[str, Any]]:
    """Parse W3H sheet. Returns list of activity dicts with keys:
    what, when, who, how, schedule, notes.
    """
    activities: list[dict[str, Any]] = []

    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, values_only=True):
        what = _normalize(row[0]) if row and len(row) > 0 else None
        if not what:
            continue
        if what.lower() in ("what", "activity"):
            continue  # header row

        activity: dict[str, Any] = {"what": what}
        if len(row) > 1:
            activity["when"] = _normalize(row[1])
        if len(row) > 2:
            activity["who"] = _normalize(row[2])
        if len(row) > 3:
            activity["how"] = _normalize(row[3])
        if len(row) > 4:
            activity["schedule"] = _normalize(row[4])
        if len(row) > 5:
            activity["notes"] = _normalize(row[5])
        activities.append(activity)

    return activities


def _parse_permission_sheet(ws: Worksheet) -> tuple[list[str], dict[str, Any]]:
    """Parse Permissions sheet.

    Returns (group_names, permission_matrix).
    group_names: list of group/role names from the header.
    permission_matrix: dict keyed by form name -> dict of privilege -> dict of group -> bool.
    """
    groups: list[str] = []
    matrix: dict[str, Any] = {}

    header_row = None
    group_start_col = 2  # default: groups start at column C (index 2)

    for row_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=ws.max_row, values_only=True), 1):
        vals = list(row)

        # Detect header
        if header_row is None:
            label0 = _normalize_lower(vals[0]) if vals else ""
            if "form" in label0 or "privilege" in label0:
                # Find where groups start
                for i, v in enumerate(vals):
                    vl = _normalize_lower(v)
                    if vl and vl not in ("form", "privileges", "privilege"):
                        group_start_col = i
                        break
                groups = [
                    _normalize(v)
                    for v in vals[group_start_col:]
                    if _normalize(v) is not None
                ]
                header_row = row_idx
                continue
            continue

        # Data rows
        form_name = _normalize(vals[0])
        privilege = _normalize(vals[1]) if len(vals) > 1 else None

        # Form name may be in previous row (merged cells pattern)
        if form_name and privilege:
            current_form = form_name
        elif not form_name and privilege:
            # Use last known form name
            pass
        else:
            continue

        if not privilege:
            continue

        if form_name:
            current_form = form_name

        if current_form not in matrix:
            matrix[current_form] = {}

        priv_grants: dict[str, bool] = {}
        for gi, group in enumerate(groups):
            col_idx = group_start_col + gi
            val = _normalize_lower(vals[col_idx]) if col_idx < len(vals) else ""
            priv_grants[group] = val in ("yes", "y", "true", "1")

        matrix[current_form][privilege] = priv_grants

    return groups, matrix


def _parse_location_hierarchy_sheet(ws: Worksheet) -> list[dict[str, Any]]:
    """Parse a Location Hierarchy sheet.

    Returns list of dicts with keys: name, level, parent (optional).
    """
    levels: list[dict[str, Any]] = []

    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, values_only=True):
        name = _normalize(row[0]) if row and len(row) > 0 else None
        if not name:
            continue
        lower = name.lower()
        if "location" in lower and ("type" in lower or "hierarchy" in lower):
            continue  # header

        # Clean up "District / Region:" style names
        cleaned = re.sub(r"\s*[/:].*$", "", name).strip()
        if cleaned:
            levels.append({"name": cleaned})

    # Assign levels (highest number = top of hierarchy)
    total = len(levels)
    for i, level in enumerate(levels):
        level["level"] = total - i
        if i > 0:
            level["parent"] = levels[i - 1]["name"]

    return levels


def _parse_programs_meta_sheet(ws: Worksheet) -> list[dict[str, Any]]:
    """Parse a Programs metadata sheet (Modelling-style).

    Returns list of dicts with keys: name, enrolment_form, exit_form.
    """
    programs: list[dict[str, Any]] = []

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
        name = _normalize(row[0]) if row and len(row) > 0 else None
        if not name:
            continue
        prog: dict[str, Any] = {"name": name}
        if len(row) > 1 and _normalize(row[1]):
            prog["enrolment_form"] = _normalize(row[1])
        if len(row) > 2 and _normalize(row[2]):
            prog["exit_form"] = _normalize(row[2])
        programs.append(prog)

    return programs


def _parse_program_encounters_sheet(ws: Worksheet) -> list[dict[str, Any]]:
    """Parse a Program Encounters metadata sheet.

    Auto-detects column layout from header row. Common layouts:
    - [Encounter Name, Subject Type, Program, Encounter Type, Frequency, ...]
    - [Encounter Name, Program, Encounter Type, ...]

    Returns list of dicts with: encounter_name, subject_type, program_name,
    encounter_type_label, cancellation_form.
    """
    mappings: list[dict[str, Any]] = []

    # Detect columns from header row
    col_map: dict[str, int] = {}
    header_row = None
    for row in ws.iter_rows(min_row=1, max_row=3, values_only=True):
        if not row:
            continue
        for i, cell in enumerate(row):
            val = _normalize(cell) if cell else ""
            lower = val.lower()
            if "encounter" in lower and "name" in lower:
                col_map["encounter_name"] = i
            elif lower == "encounter name" or (i == 0 and "encounter" in lower):
                col_map["encounter_name"] = i
            elif "subject" in lower and "type" in lower:
                col_map["subject_type"] = i
            elif lower in ("program", "program name"):
                col_map["program_name"] = i
            elif "encounter type" in lower and "name" not in lower:
                col_map["encounter_type_label"] = i
            elif "frequency" in lower:
                col_map["frequency"] = i
            elif "cancel" in lower:
                col_map["cancellation_form"] = i
        if col_map:
            header_row = row
            break

    # Fallback: assume standard layout if no header detected
    if "encounter_name" not in col_map:
        col_map = {"encounter_name": 0, "subject_type": 1, "program_name": 2,
                    "encounter_type_label": 3, "frequency": 4, "cancellation_form": 9}

    start_row = 3 if header_row else 2
    for row in ws.iter_rows(min_row=start_row, max_row=ws.max_row, values_only=True):
        enc_idx = col_map.get("encounter_name", 0)
        enc_name = _normalize(row[enc_idx]) if row and len(row) > enc_idx and row[enc_idx] else None
        if not enc_name:
            continue
        mapping: dict[str, Any] = {"encounter_name": enc_name}
        if "subject_type" in col_map and col_map["subject_type"] < len(row):
            val = _normalize(row[col_map["subject_type"]])
            if val:
                mapping["subject_type"] = val
        if "program_name" in col_map and col_map["program_name"] < len(row):
            val = _normalize(row[col_map["program_name"]])
            if val:
                mapping["program_name"] = val
        if "cancellation_form" in col_map and col_map["cancellation_form"] < len(row):
            val = _normalize(row[col_map["cancellation_form"]])
            if val:
                mapping["cancellation_form"] = val
        mappings.append(mapping)

    return mappings


def _parse_form_summary_sheet(ws: Worksheet) -> list[dict[str, Any]]:
    """Parse a Form Summary sheet — the single source of truth for form→program→encounter mapping.

    Expected columns: Sr. No., Form, Subject Type, Program, Encounter Type, ...
    Stops at "Out of Scope" marker row (everything after is unimplemented).

    Returns list of dicts with: form_name, subject_type, program, encounter_type, in_scope.
    """
    entries: list[dict[str, Any]] = []

    # Detect header row and column mapping
    col_map: dict[str, int] = {}
    header_row: int | None = None
    _col_patterns = {
        "sr_no": re.compile(r"sr\.?\s*no|serial|#", re.I),
        "form_name": re.compile(r"^form$|form\s*name", re.I),
        "subject_type": re.compile(r"subject\s*type", re.I),
        "program": re.compile(r"^program$|program\s*name", re.I),
        "encounter_type": re.compile(r"encounter\s*type", re.I),
    }

    for row_idx, row in enumerate(
        ws.iter_rows(min_row=1, max_row=min(10, ws.max_row), values_only=True), 1
    ):
        if not row:
            continue
        matched = 0
        temp_map: dict[str, int] = {}
        for col_idx, cell in enumerate(row):
            if cell is None:
                continue
            cell_str = str(cell).strip()
            for key, pat in _col_patterns.items():
                if pat.search(cell_str) and key not in temp_map:
                    temp_map[key] = col_idx
                    matched += 1
        if matched >= 3:  # Need at least form_name + program + encounter_type
            header_row = row_idx
            col_map = temp_map
            break

    if header_row is None or "form_name" not in col_map:
        logger.warning("Form Summary sheet: could not detect header row")
        return entries

    logger.info(
        "Form Summary sheet: header at row %d, columns: %s",
        header_row, {k: v for k, v in col_map.items()},
    )

    in_scope = True
    for row in ws.iter_rows(min_row=header_row + 1, max_row=ws.max_row, values_only=True):
        if not row:
            continue

        # Check for "Out of Scope" marker
        first_val = _normalize(row[0]) if len(row) > 0 else None
        if first_val and "out of scope" in first_val.lower():
            in_scope = False
            continue

        form_name = _normalize(row[col_map["form_name"]]) if col_map.get("form_name") is not None and len(row) > col_map["form_name"] else None
        if not form_name:
            continue

        entry: dict[str, Any] = {
            "form_name": form_name,
            "in_scope": in_scope,
        }

        if "subject_type" in col_map and len(row) > col_map["subject_type"]:
            entry["subject_type"] = _normalize(row[col_map["subject_type"]]) or "Individual"

        if "program" in col_map and len(row) > col_map["program"]:
            entry["program"] = _normalize(row[col_map["program"]]) or ""

        if "encounter_type" in col_map and len(row) > col_map["encounter_type"]:
            et = _normalize(row[col_map["encounter_type"]]) or ""
            entry["encounter_type"] = et if et.upper() != "NA" else ""

        entries.append(entry)

    logger.info(
        "Form Summary: %d entries (%d in scope, %d out of scope)",
        len(entries),
        sum(1 for e in entries if e["in_scope"]),
        sum(1 for e in entries if not e["in_scope"]),
    )
    return entries


def _parse_visit_scheduling_sheet(ws: Worksheet) -> list[dict[str, Any]]:
    """Parse a Visit Scheduling sheet into structured scheduling data.

    Expected columns (flexible matching):
    - On Completion Of / Trigger Form: which form/encounter triggers the schedule
    - Schedule Form / Visit Type: what encounter gets scheduled
    - Frequency / Repeats: how often (once, monthly, weekly, etc.)
    - Due Date / Schedule Days / Days After: days until due
    - Overdue Date / Overdue Days: days until overdue
    - Cancellation / On Cancel: what happens on cancel (reschedule, close, etc.)
    - Conditions / When / Criteria: conditions for scheduling

    Returns list of dicts with: trigger, schedule_encounter, due_days, overdue_days,
    frequency, cancel_behavior, conditions.
    """
    schedules: list[dict[str, Any]] = []

    # Find header row and map columns
    header_row: int | None = None
    col_map: dict[str, int] = {}

    _col_patterns = {
        "trigger": re.compile(r"on\s+completion|trigger|after|source\s+form", re.I),
        "schedule_encounter": re.compile(r"schedule\s+form|visit\s+type|schedule\s+encounter|next\s+visit|target", re.I),
        "frequency": re.compile(r"frequen|repeat|recur", re.I),
        "due_days": re.compile(r"due\s+date|schedule\s+day|days?\s+after|due\s+day|days?\s+to\s+due", re.I),
        "overdue_days": re.compile(r"overdue|max\s+day|days?\s+to\s+overdue", re.I),
        "cancel_behavior": re.compile(r"cancel|on\s+cancel|cancellation", re.I),
        "conditions": re.compile(r"condition|when|criteria|rule|logic", re.I),
    }

    for row_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=min(10, ws.max_row), values_only=True), 1):
        if not row:
            continue
        matched = 0
        temp_map: dict[str, int] = {}
        for col_idx, cell in enumerate(row):
            if cell is None:
                continue
            cell_str = str(cell).strip()
            for key, pat in _col_patterns.items():
                if pat.search(cell_str) and key not in temp_map:
                    temp_map[key] = col_idx
                    matched += 1
        if matched >= 2:
            header_row = row_idx
            col_map = temp_map
            break

    if header_row is None:
        logger.warning("Visit Scheduling sheet: could not detect header row")
        return schedules

    logger.info(
        "Visit Scheduling sheet: header at row %d, columns: %s",
        header_row, {k: v for k, v in col_map.items()},
    )

    # Parse data rows
    for row in ws.iter_rows(min_row=header_row + 1, max_row=ws.max_row, values_only=True):
        if not row or all(c is None for c in row):
            continue

        def _get(key: str) -> str | None:
            idx = col_map.get(key)
            if idx is None or idx >= len(row):
                return None
            return _normalize(row[idx])

        trigger = _get("trigger")
        schedule_enc = _get("schedule_encounter")
        if not trigger and not schedule_enc:
            continue

        entry: dict[str, Any] = {}
        if trigger:
            entry["trigger"] = trigger
        if schedule_enc:
            entry["schedule_encounter"] = schedule_enc

        # Parse due days (could be "28 days", "4 weeks", or just "28")
        due_raw = _get("due_days")
        if due_raw:
            days = _parse_days_value(due_raw)
            if days is not None:
                entry["due_days"] = days
            else:
                entry["due_days_raw"] = due_raw

        overdue_raw = _get("overdue_days")
        if overdue_raw:
            days = _parse_days_value(overdue_raw)
            if days is not None:
                entry["overdue_days"] = days
            else:
                entry["overdue_days_raw"] = overdue_raw

        frequency = _get("frequency")
        if frequency:
            entry["frequency"] = frequency

        cancel = _get("cancel_behavior")
        if cancel:
            entry["cancel_behavior"] = cancel

        conditions = _get("conditions")
        if conditions:
            entry["conditions"] = conditions

        schedules.append(entry)

    logger.info("Visit Scheduling: parsed %d schedule entries", len(schedules))
    return schedules


def _parse_days_value(raw: str) -> int | None:
    """Parse a days value from various formats: '28', '28 days', '4 weeks', '1 month'."""
    raw = raw.strip().lower()

    # Direct number
    m = re.match(r"^(\d+)\s*(?:days?)?$", raw)
    if m:
        return int(m.group(1))

    # Weeks
    m = re.match(r"^(\d+)\s*weeks?$", raw)
    if m:
        return int(m.group(1)) * 7

    # Months (approximate)
    m = re.match(r"^(\d+)\s*months?$", raw)
    if m:
        return int(m.group(1)) * 30

    return None


def _parse_modelling_sheet(ws: Worksheet) -> dict[str, Any]:
    """Parse a Modelling sheet.

    Returns dict with: subject_types, programs, encounters.
    """
    subject_types: list[dict[str, Any]] = []
    programs: list[dict[str, Any]] = []
    encounters: list[dict[str, Any]] = []

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
        row_type = _normalize_lower(row[0]) if row and len(row) > 0 else ""
        name = _normalize(row[1]) if len(row) > 1 else None
        if not name:
            continue

        if row_type == "subject":
            st: dict[str, Any] = {"name": name}
            if len(row) > 2 and _normalize(row[2]):
                st["type"] = _normalize(row[2])
            subject_types.append(st)
        elif row_type == "program":
            prog: dict[str, Any] = {"name": name}
            if len(row) > 4 and _normalize(row[4]):
                prog["colour"] = _normalize(row[4])
            programs.append(prog)
        elif row_type == "encounter":
            enc: dict[str, Any] = {"name": name}
            if len(row) > 2 and _normalize(row[2]):
                enc["subject_type"] = _normalize(row[2])
            if len(row) > 6 and _normalize(row[6]):
                enc["program"] = _normalize(row[6])
            if len(row) > 7 and _normalize(row[7]):
                enc["encounter_type"] = _normalize(row[7])
            encounters.append(enc)

    return {
        "subject_types": subject_types,
        "programs": programs,
        "encounters": encounters,
    }


def _parse_form_sheet(
    ws: Worksheet,
    sheet_name: str,
) -> SRSFormDefinition | None:
    """Parse a single form sheet into an SRSFormDefinition.

    Auto-detects the header row by looking for "Field Name" and "Data Type".
    Groups fields by the Page/Group Name column.
    """
    # Find header row
    header_row_idx: int | None = None
    header_row_data: tuple | None = None

    for row_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=min(5, ws.max_row), values_only=True), 1):
        if _is_form_header_row(row):
            header_row_idx = row_idx
            header_row_data = row
            break

    if header_row_idx is None or header_row_data is None:
        logger.debug("Sheet '%s' has no form header row, skipping", sheet_name)
        return None

    cols = _detect_column_indices(header_row_data)
    if "field_name" not in cols or "data_type" not in cols:
        logger.debug("Sheet '%s' missing field_name/data_type columns, skipping", sheet_name)
        return None

    # Parse data rows
    groups_ordered: list[str] = []
    groups_fields: dict[str, list[SRSFormField]] = {}
    current_group = "Default"
    skip_row_patterns = {"tb patients form"}

    for row in ws.iter_rows(min_row=header_row_idx + 1, max_row=ws.max_row, values_only=True):
        vals = list(row)

        # Extract page/group name
        page_val = _normalize(vals[cols["page_name"]]) if "page_name" in cols and cols["page_name"] < len(vals) else None
        field_name = _normalize(vals[cols["field_name"]]) if cols["field_name"] < len(vals) else None
        data_type_raw = _normalize(vals[cols["data_type"]]) if cols["data_type"] < len(vals) else None

        # Update current group from page name column
        if page_val:
            current_group = page_val

        # Skip rows with no field name
        if not field_name:
            continue

        # Skip known noise rows (sub-headers within forms)
        if field_name.lower() in skip_row_patterns:
            continue
        # Skip rows where "field name" looks like a sub-heading (no data type)
        if not data_type_raw:
            # Sometimes the field name IS a section header if data_type is missing
            # But only skip if it looks like a title
            if page_val and page_val.lower() == field_name.lower():
                continue
            # Field with no data type -- skip
            continue

        # Map data type
        dt_lower = data_type_raw.lower().strip()
        avni_type = _DATA_TYPE_MAP.get(dt_lower, "Text")

        # Determine if mandatory
        mandatory = False
        if "mandatory" in cols and cols["mandatory"] < len(vals):
            mandatory = _is_yes(vals[cols["mandatory"]])

        # Determine selection type (SingleSelect / MultiSelect)
        selection_type: str | None = None
        if "selection_type" in cols and cols["selection_type"] < len(vals):
            sel = _normalize_lower(vals[cols["selection_type"]])
            if "multi" in sel:
                selection_type = "MultiSelect"
            elif "single" in sel:
                selection_type = "SingleSelect"

        # Infer selection type from data type if not set
        if avni_type == "Coded" and selection_type is None:
            if "multi" in dt_lower:
                selection_type = "MultiSelect"
            else:
                selection_type = "SingleSelect"

        # Parse options
        options: list[str] = []
        if "options" in cols and cols["options"] < len(vals):
            raw_opt = _normalize(vals[cols["options"]])
            # Guard: if the OPTIONS column contains a selection type value
            # (e.g. "Single Select", "Multi Select"), it's misaligned data.
            # In that case, read the selection type from this column and
            # look one column to the right for the actual options.
            if raw_opt and raw_opt.lower().strip() in (
                "single select", "multi select", "singleselect", "multiselect",
                "single", "multi",
            ):
                if "multi" in raw_opt.lower():
                    selection_type = "MultiSelect"
                else:
                    selection_type = "SingleSelect"
                # Try the next column for actual options
                next_col = cols["options"] + 1
                if next_col < len(vals):
                    options = _parse_options(vals[next_col])
            else:
                options = _parse_options(raw_opt)

        # If no options found in the options column but the selection_type column
        # contains what look like actual option values (Gubbachi misalignment)
        if not options and avni_type == "Coded":
            if "selection_type" in cols and cols["selection_type"] < len(vals):
                sel_raw = _normalize(vals[cols["selection_type"]])
                if sel_raw and sel_raw.lower().strip() not in (
                    "single select", "multi select", "singleselect", "multiselect",
                    "single", "multi", "",
                ):
                    # The selection_type column has actual option values
                    candidate_opts = _parse_options(sel_raw)
                    if len(candidate_opts) >= 2:
                        options = candidate_opts

        # Parse unit
        unit: str | None = None
        if "unit" in cols and cols["unit"] < len(vals):
            unit = _parse_unit(vals[cols["unit"]])

        # Extract unit from field name if not set
        if not unit and avni_type == "Numeric":
            field_name, extracted_unit = _extract_unit_from_name(field_name)
            if extracted_unit:
                unit = extracted_unit

        # Parse min/max
        low: float | None = None
        high: float | None = None
        if "min_max" in cols and cols["min_max"] < len(vals):
            low, high = _parse_min_max(vals[cols["min_max"]])
        # Also check validation column
        if low is None and high is None and "validation" in cols and cols["validation"] < len(vals):
            low, high = _parse_min_max(vals[cols["validation"]])

        # Build key values for skip logic
        key_values: list[dict[str, Any]] = []

        # Show when / hide when
        show_when = None
        hide_when = None
        if "show_when" in cols and cols["show_when"] < len(vals):
            show_when = _normalize(vals[cols["show_when"]])
        if "hide_when" in cols and cols["hide_when"] < len(vals):
            hide_when = _normalize(vals[cols["hide_when"]])

        if show_when:
            key_values.append({"key": "showWhen", "value": show_when})
        if hide_when:
            key_values.append({"key": "hideWhen", "value": hide_when})

        # Unique option for multi-select
        if "unique_option" in cols and cols["unique_option"] < len(vals):
            unique_opt = _normalize(vals[cols["unique_option"]])
            if unique_opt:
                key_values.append({"key": "ExcludedAnswers", "value": unique_opt})

        # Build field
        field = SRSFormField(
            name=field_name,
            dataType=avni_type,
            mandatory=mandatory,
            options=options if options else None,
            type=selection_type,
            unit=unit,
            lowAbsolute=low,
            highAbsolute=high,
            keyValues=key_values if key_values else None,
        )

        # Add to group
        if current_group not in groups_fields:
            groups_ordered.append(current_group)
            groups_fields[current_group] = []
        groups_fields[current_group].append(field)

    # Build form definition
    if not groups_fields:
        logger.debug("Sheet '%s' has no parseable fields, skipping", sheet_name)
        return None

    form_groups = [
        SRSFormGroup(name=gname, fields=groups_fields[gname])
        for gname in groups_ordered
    ]

    return SRSFormDefinition(
        name=sheet_name.strip(),
        formType="ProgramEncounter",  # Will be refined later
        groups=form_groups,
    )


# ---------------------------------------------------------------------------
# Main Parser
# ---------------------------------------------------------------------------


class SRSParser:
    """Parse Avni SRS Excel files into structured SRSData."""

    def __init__(self, file_path: str | Path) -> None:
        self.file_path = Path(file_path)
        self.wb = openpyxl.load_workbook(str(file_path), data_only=True)
        self.sheet_names = self.wb.sheetnames

        # Sector hint (set externally before parse() if known)
        self._sector: str | None = None

        # Intermediate state
        self._summary: dict[str, Any] = {}
        self._user_groups: list[str] = []
        self._w3h: list[dict[str, Any]] = []
        self._forms: list[SRSFormDefinition] = []
        self._permission_groups: list[str] = []
        self._permission_matrix: dict[str, Any] = {}
        self._location_hierarchy: list[dict[str, Any]] = []
        self._modelling: dict[str, Any] = {}
        self._programs_meta: list[dict[str, Any]] = []
        self._program_encounters_meta: list[dict[str, Any]] = []
        self._encounters_meta: list[dict[str, Any]] = []
        self._visit_schedules: list[dict[str, Any]] = []
        self._form_summary: list[dict[str, Any]] = []  # From "Form Summary" sheet

    def parse(self) -> SRSData:
        """Parse the complete SRS workbook into SRSData."""
        logger.info("Parsing SRS file: %s (%d sheets)", self.file_path.name, len(self.sheet_names))

        # Classify and parse each sheet
        for name in self.sheet_names:
            ws = self.wb[name]
            classification = _classify_sheet(name)
            logger.debug("Sheet '%s' classified as '%s'", name, classification)

            if classification == "skip":
                continue
            elif classification == "form_summary":
                self._form_summary = _parse_form_summary_sheet(ws)
            elif classification == "summary":
                self._summary = _parse_summary_sheet(ws)
            elif classification == "user":
                self._user_groups = _parse_user_sheet(ws)
            elif classification == "w3h":
                self._w3h = _parse_w3h_sheet(ws)
            elif classification == "permission":
                self._permission_groups, self._permission_matrix = _parse_permission_sheet(ws)
            elif classification == "location":
                self._location_hierarchy = _parse_location_hierarchy_sheet(ws)
            elif classification == "programs_meta":
                self._programs_meta = _parse_programs_meta_sheet(ws)
            elif classification == "visit_scheduling":
                self._visit_schedules = _parse_visit_scheduling_sheet(ws)
            elif classification == "program_encounters":
                self._program_encounters_meta = _parse_program_encounters_sheet(ws)
            elif classification == "encounters_meta":
                self._encounters_meta = []
                for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
                    enc_name = _normalize(row[0]) if row and len(row) > 0 else None
                    if enc_name:
                        entry: dict[str, Any] = {"name": enc_name}
                        # Column 1 is often Subject Type
                        if len(row) > 1 and _normalize(row[1]):
                            entry["subject_type"] = _normalize(row[1])
                        self._encounters_meta.append(entry)
            elif classification == "subject_types":
                # Separate Subject Types sheet (from modelling docs)
                for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
                    st_name = _normalize(row[0]) if row and len(row) > 0 else None
                    if st_name:
                        st_type = _normalize(row[1]) if len(row) > 1 and row[1] else None
                        entry: dict[str, Any] = {"name": st_name}
                        if st_type:
                            entry["type"] = st_type
                        if not self._modelling.get("subject_types"):
                            self._modelling["subject_types"] = []
                        self._modelling["subject_types"].append(entry)
            elif classification == "modelling":
                self._modelling = _parse_modelling_sheet(ws)
            elif classification == "report" or classification == "dashboard":
                # Informational only -- skip for bundle generation
                continue
            else:
                # Unknown sheet -- try to parse as a form
                form_def = _parse_form_sheet(ws, name)
                if form_def:
                    self._forms.append(form_def)
                    logger.info(
                        "Parsed form sheet '%s': %d groups, %d fields",
                        name,
                        len(form_def.groups),
                        sum(len(g.fields) for g in form_def.groups),
                    )

        # Assemble SRSData
        return self._build_srs_data()

    def _build_srs_data(self) -> SRSData:
        """Assemble parsed data into an SRSData instance."""
        # Organization name
        org_name = (
            self._summary.get("org_name")
            or self._summary.get("program_name")
            or self.file_path.stem
        )

        # Subject types
        subject_types: list[dict[str, Any]] = []
        if self._modelling.get("subject_types"):
            for st in self._modelling["subject_types"]:
                subject_types.append({
                    "name": st["name"],
                    "type": st.get("type", "Person"),
                })
        if not subject_types:
            subject_types = self._infer_subject_types_from_forms()
        if not subject_types:
            subject_types = [{"name": "Individual", "type": "Person"}]

        # Programs -- from modelling, programs_meta, or infer from W3H/form names
        programs = self._resolve_programs()

        # Encounter types -- collected from form classification
        encounter_types = self._resolve_encounter_types()

        # Classify forms (set correct formType, programName, encounterTypeName)
        self._classify_forms(programs, encounter_types)

        # Sector-aware classification (uses production patterns to fix remaining issues)
        try:
            from app.services.sector_classifier import classify_forms_by_sector
            sector = self._summary.get("sector") or self._sector
            classify_forms_by_sector(
                forms=self._forms,
                subject_types=subject_types,
                programs=programs,
                program_encounters_meta=self._program_encounters_meta,
                encounters_meta=self._encounters_meta,
                sector=sector,
            )
        except Exception as e:
            logger.warning("Sector classification failed (non-fatal): %s", e)

        # All unique encounter type names
        all_encounter_names: list[str] = []
        seen_enc: set[str] = set()
        for form_def in self._forms:
            if form_def.encounterTypeName and form_def.encounterTypeName not in seen_enc:
                all_encounter_names.append(form_def.encounterTypeName)
                seen_enc.add(form_def.encounterTypeName)

        # Groups
        groups = self._resolve_groups()

        # Address level types
        address_level_types = self._resolve_address_levels()

        # Program-encounter mappings
        pe_mappings = self._resolve_program_encounter_mappings()

        # General encounter types (no program)
        general_encounter_types = self._resolve_general_encounter_types()

        return SRSData(
            orgName=org_name,
            subjectTypes=subject_types,
            programs=[{"name": p} if isinstance(p, str) else p for p in programs],
            encounterTypes=all_encounter_names,
            forms=self._forms,
            groups=groups,
            addressLevelTypes=address_level_types if address_level_types else None,
            programEncounterMappings=pe_mappings if pe_mappings else None,
            generalEncounterTypes=general_encounter_types if general_encounter_types else None,
            visitSchedules=self._visit_schedules if self._visit_schedules else None,
        )

    def _resolve_programs(self) -> list[dict[str, Any]]:
        """Determine program list from available data sources."""
        programs: list[dict[str, Any]] = []
        seen: set[str] = set()

        # From Form Summary sheet (highest priority — has canonical names)
        if self._form_summary:
            for entry in self._form_summary:
                if not entry.get("in_scope", True):
                    continue
                prog = (entry.get("program") or "").strip()
                if prog and prog not in seen:
                    programs.append({"name": prog})
                    seen.add(prog)

        # From modelling sheet
        for p in self._modelling.get("programs", []):
            name = p["name"].strip()
            if name not in seen:
                prog_entry: dict[str, Any] = {"name": name}
                if p.get("colour"):
                    prog_entry["colour"] = p["colour"]
                programs.append(prog_entry)
                seen.add(name)

        # From programs meta sheet
        for p in self._programs_meta:
            name = p["name"].strip()
            if name not in seen:
                programs.append({"name": name})
                seen.add(name)

        # Infer from W3H (enrollment/exit activities suggest programs)
        if not programs:
            programs = self._infer_programs_from_sheets()

        return programs

    def _infer_programs_from_sheets(self) -> list[dict[str, Any]]:
        """Infer programs from form sheet names.

        Heuristic: sheets named "X Enrollment" suggest a program called "X Program"
        or similar.
        """
        programs: list[dict[str, Any]] = []
        seen: set[str] = set()

        enrollment_pattern = re.compile(
            r"^(.+?)\s*(enroll?ment|enrol)\s*$", re.IGNORECASE
        )

        for name in self.sheet_names:
            m = enrollment_pattern.match(name.strip())
            if m:
                prog_base = m.group(1).strip()
                # Check if there's a matching exit sheet
                prog_name = prog_base
                if prog_name not in seen:
                    programs.append({"name": prog_name})
                    seen.add(prog_name)

        return programs

    def _infer_subject_types_from_forms(self) -> list[dict[str, Any]]:
        """Infer subject types from registration form sheet names.

        Heuristic: sheets named "X Reg" or "X Registration" where X is NOT
        "Individual" or "Family" suggest a non-person subject type (e.g. "Kiln").
        We also always include "Individual" if there is any person-like registration.
        """
        subject_types: list[dict[str, Any]] = []
        seen: set[str] = set()

        # Person-like registration keywords
        person_keywords = {
            "individual", "worker", "person", "member", "beneficiary",
            "patient", "child", "mother", "woman", "women", "student",
            "family", "household",
        }

        reg_pattern = re.compile(
            r"^(.+?)\s*(registration|reg)\s*$", re.IGNORECASE
        )

        has_person_reg = False

        for form_def in self._forms:
            m = reg_pattern.match(form_def.name.strip())
            if not m:
                continue

            base = m.group(1).strip()
            base_lower = base.lower()

            # Check if this is a person-like registration
            if any(kw in base_lower for kw in person_keywords):
                has_person_reg = True
                continue

            # Non-person subject type (e.g. "Kiln Reg" -> subject type "Kiln")
            if base_lower not in seen:
                subject_types.append({"name": base, "type": "Individual"})
                seen.add(base_lower)
                # Update the form to reference this subject type
                form_def.subjectTypeName = base

        # Always include Individual if there's a person-like registration
        # or if there are enrollment forms (which typically act on Individual subjects)
        if has_person_reg or any(
            f.formType == "ProgramEnrolment" for f in self._forms
        ):
            if "individual" not in seen:
                subject_types.insert(0, {"name": "Individual", "type": "Person"})
                seen.add("individual")

        return subject_types

    def _resolve_encounter_types(self) -> dict[str, str]:
        """Build encounter_name -> form_type mapping.

        Returns dict: encounter_type_name -> Avni form type.
        Uses Form Summary as primary source (has canonical encounter type names).
        """
        encounters: dict[str, str] = {}

        # From Form Summary sheet (highest priority)
        if self._form_summary:
            for entry in self._form_summary:
                if not entry.get("in_scope", True):
                    continue
                et = (entry.get("encounter_type") or "").strip()
                prog = (entry.get("program") or "").strip()
                if et:
                    if prog:
                        encounters[et] = "ProgramEncounter"
                    else:
                        encounters[et] = "Encounter"

        # From modelling sheet
        for enc in self._modelling.get("encounters", []):
            name = enc["name"]
            if name not in encounters:
                etype = enc.get("encounter_type", "ProgramEncounter")
                encounters[name] = etype

        # From program encounters meta
        for pe in self._program_encounters_meta:
            name = pe["encounter_name"]
            if name not in encounters:
                encounters[name] = "ProgramEncounter"

        # From general encounters meta
        for enc in self._encounters_meta:
            name = enc["name"]
            if name not in encounters:
                encounters[name] = "Encounter"

        return encounters

    def _match_form_to_summary(
        self,
        form_name: str,
    ) -> dict[str, Any] | None:
        """Match a parsed form (possibly with truncated sheet name) to a Form Summary entry.

        Excel sheet names are limited to 31 chars, so we match by:
        1. Exact match (case-insensitive)
        2. Form Summary name starts with the sheet name (truncation)
        3. Sheet name starts with the Form Summary name
        4. Key words overlap (handles "Child PNC" vs "Child PNC (Neonatal)",
           "Children Monthly Gradation" vs "Child Monthly Gradation")
        """
        if not self._form_summary:
            return None

        name_lower = form_name.strip().lower()
        # Only consider in-scope entries for matching
        in_scope = [e for e in self._form_summary if e.get("in_scope", True)]

        # Pass 1: Exact match
        for entry in in_scope:
            fs_name = entry["form_name"].strip().lower()
            if name_lower == fs_name:
                return entry

        # Pass 2: Truncation — sheet name is prefix of full name
        for entry in in_scope:
            fs_name = entry["form_name"].strip().lower()
            if len(name_lower) >= 20 and fs_name.startswith(name_lower):
                return entry
            if len(fs_name) >= 20 and name_lower.startswith(fs_name):
                return entry

        # Pass 3: One is prefix of the other (short names like "ANC", "Delivery")
        for entry in in_scope:
            fs_name = entry["form_name"].strip().lower()
            if name_lower.startswith(fs_name) or fs_name.startswith(name_lower):
                return entry

        # Pass 4: Core word overlap — handles variant naming
        # e.g., "Child PNC" matches "Child PNC (Neonatal)",
        #        "Child Monthly Gradation Form Fo" matches "Children Monthly Gradation Form"
        # Strip minimal filler words and compare meaningful tokens
        _filler = {"", "the", "for", "of", "a"}
        name_words = set(re.split(r"[\s\-_()]+", name_lower)) - _filler
        best_match = None
        best_score = 0
        for entry in in_scope:
            fs_name = entry["form_name"].strip().lower()
            fs_words = set(re.split(r"[\s\-_()]+", fs_name)) - _filler
            if not fs_words or not name_words:
                continue
            overlap_words = name_words & fs_words
            overlap = len(overlap_words)
            union = len(name_words | fs_words)
            # Jaccard similarity with bonus for first-word match
            base_score = overlap / union if union > 0 else 0
            # Bonus: if the first meaningful word of the sheet name matches,
            # add 0.1 to prefer "Malnutrition Exit Form" over "Pregnancy Program Exit"
            # when matching "Malnutrition Program Exit"
            first_word = name_lower.split()[0] if name_lower.split() else ""
            first_word_bonus = 0.1 if first_word in fs_words else 0
            score = base_score + first_word_bonus
            # Require at least 40% Jaccard overlap and at least 2 shared words
            if score > best_score and base_score >= 0.4 and overlap >= 2:
                best_score = score
                best_match = entry

        return best_match

    def _classify_forms(
        self,
        programs: list[dict[str, Any]],
        encounter_types: dict[str, str],
    ) -> None:
        """Assign formType, programName, encounterTypeName to each parsed form.

        Uses Form Summary sheet as the primary source of truth (if available).
        Falls back to heuristic classification for forms not in Form Summary.
        """
        program_names = {p["name"].strip().lower(): p["name"].strip() for p in programs}

        # ── PRIMARY: Use Form Summary sheet if available ──
        if self._form_summary:
            logger.info("Using Form Summary sheet for form classification (%d entries)", len(self._form_summary))

            # Filter: only keep forms that match an in-scope Form Summary entry
            # Forms that don't match any in-scope entry are considered out-of-scope
            matched_forms: list[SRSFormDefinition] = []
            unmatched_forms: list[str] = []
            for f in self._forms:
                match = self._match_form_to_summary(f.name)
                if match:
                    matched_forms.append(f)
                else:
                    unmatched_forms.append(f.name)

            if unmatched_forms:
                logger.info(
                    "Removed %d forms not found in Form Summary (likely out of scope): %s",
                    len(unmatched_forms), unmatched_forms,
                )
            self._forms = matched_forms

            # Classify each form using Form Summary
            for form_def in self._forms:
                match = self._match_form_to_summary(form_def.name)
                if not match:
                    # No Form Summary match — will fall through to heuristic below
                    continue

                program = (match.get("program") or "").strip()
                encounter_type = (match.get("encounter_type") or "").strip()
                form_name_full = match["form_name"].strip()

                # Use full name from Form Summary (fixes truncation)
                if len(form_name_full) > len(form_def.name.strip()):
                    logger.info(
                        "Expanding truncated form name: '%s' -> '%s'",
                        form_def.name, form_name_full,
                    )
                    form_def.name = form_name_full

                # Determine formType based on program + encounter_type
                name_lower = form_def.name.strip().lower()

                if "registration" in name_lower or "reg" == name_lower.split()[-1]:
                    form_def.formType = "IndividualProfile"
                elif "enrolment" in name_lower or "enrollment" in name_lower:
                    form_def.formType = "ProgramEnrolment"
                    form_def.programName = program
                elif "exit" in name_lower:
                    form_def.formType = "ProgramExit"
                    form_def.programName = program
                elif program and encounter_type:
                    # Has both program and encounter type → ProgramEncounter
                    form_def.formType = "ProgramEncounter"
                    form_def.programName = program
                    form_def.encounterTypeName = encounter_type
                elif encounter_type and not program:
                    # Encounter type but no program → general Encounter
                    form_def.formType = "Encounter"
                    form_def.encounterTypeName = encounter_type
                elif program and not encounter_type:
                    # Program but no encounter type → likely enrolment
                    form_def.formType = "ProgramEnrolment"
                    form_def.programName = program
                else:
                    # No program, no encounter type — use heuristic below
                    pass

                if form_def.formType:
                    logger.debug(
                        "Form Summary classified '%s': formType=%s, program=%s, encounter=%s",
                        form_def.name, form_def.formType,
                        form_def.programName, form_def.encounterTypeName,
                    )

            # For forms that got classified via Form Summary, skip heuristic
            # For unclassified ones, fall through to heuristic below

        # ── FALLBACK: Heuristic classification for forms not yet classified ──
        # Build program encounter mapping from meta
        pe_map: dict[str, str] = {}  # encounter_name -> program_name
        for pe in self._program_encounters_meta:
            enc = pe["encounter_name"].strip()
            prog = pe.get("program_name", "").strip()
            if prog:
                pe_map[enc.lower()] = prog
        for enc in self._modelling.get("encounters", []):
            prog = enc.get("program", "")
            if prog:
                pe_map[enc["name"].strip().lower()] = prog

        # Also add pe_map entries from Form Summary (for heuristic cancellation matching)
        for entry in self._form_summary:
            prog = (entry.get("program") or "").strip()
            et = (entry.get("encounter_type") or "").strip()
            if prog and et:
                pe_map[et.lower()] = prog

        # Build program enrolment/exit form names from meta
        prog_enrol_forms: dict[str, str] = {}  # form_name_lower -> program_name
        prog_exit_forms: dict[str, str] = {}
        for pm in self._programs_meta:
            pname = pm["name"].strip()
            if pm.get("enrolment_form"):
                prog_enrol_forms[pm["enrolment_form"].strip().lower()] = pname
            if pm.get("exit_form"):
                prog_exit_forms[pm["exit_form"].strip().lower()] = pname

        for form_def in self._forms:
            # Skip forms already classified by Form Summary
            if form_def.formType and form_def.formType != "ProgramEncounter":
                continue
            if form_def.formType == "ProgramEncounter" and form_def.programName:
                continue
            name_lower = form_def.name.strip().lower()

            # 1. Registration forms (match "registration", "reg", ending with " reg")
            is_reg = (
                "registration" in name_lower
                or re.search(r"\breg\b", name_lower)
                or name_lower.endswith(" reg")
            )
            if is_reg and "enrol" not in name_lower and "enrollment" not in name_lower:
                form_def.formType = "IndividualProfile"
                continue

            # 2. Check enrolment form from meta
            if name_lower in prog_enrol_forms:
                form_def.formType = "ProgramEnrolment"
                form_def.programName = prog_enrol_forms[name_lower]
                continue

            # 3. Check exit form from meta
            if name_lower in prog_exit_forms:
                form_def.formType = "ProgramExit"
                form_def.programName = prog_exit_forms[name_lower]
                continue

            # 4. Check if encounter type from modelling/meta
            if name_lower in encounter_types:
                etype = encounter_types[name_lower]
                form_def.formType = etype
                form_def.encounterTypeName = form_def.name.strip()
                if name_lower in pe_map:
                    form_def.programName = pe_map[name_lower]
                continue

            # 5. Heuristic: "X Enrollment" / "X Enrolment" -> ProgramEnrolment
            enrol_match = re.match(r"^(.+?)\s*(enroll?ment|enrol)\s*$", name_lower)
            if enrol_match:
                base = enrol_match.group(1).strip()
                form_def.formType = "ProgramEnrolment"
                # Try to match to a program
                form_def.programName = self._find_matching_program(base, program_names)
                continue

            # 6. Heuristic: "X Exit" -> ProgramExit
            exit_match = re.match(r"^(.+?)\s*exit\s*$", name_lower)
            if exit_match:
                base = exit_match.group(1).strip()
                form_def.formType = "ProgramExit"
                form_def.programName = self._find_matching_program(base, program_names)
                continue

            # 7. Heuristic: "X Follow Up" / "X Screening" / "X Check Up" -> ProgramEncounter
            encounter_match = re.match(
                r"^(.+?)\s*(follow\s*up|screening|check\s*up|health\s*check)\s*$",
                name_lower,
            )
            if encounter_match:
                form_def.formType = "ProgramEncounter"
                form_def.encounterTypeName = form_def.name.strip()
                # Try to find which program this belongs to
                base = encounter_match.group(1).strip()
                if name_lower in pe_map:
                    form_def.programName = pe_map[name_lower]
                else:
                    form_def.programName = self._find_matching_program(base, program_names)
                continue

            # 8. Heuristic: "X Cancellation" -> ProgramEncounterCancellation or IndividualEncounterCancellation
            cancel_match = re.match(r"^(.+?)\s*cancell?ation\s*$", name_lower)
            if cancel_match:
                base_name = cancel_match.group(1).strip()
                # Check if there's a program encounter with this base name
                if base_name.lower() in pe_map:
                    form_def.formType = "ProgramEncounterCancellation"
                    form_def.encounterTypeName = base_name
                    form_def.programName = pe_map[base_name.lower()]
                else:
                    form_def.formType = "IndividualEncounterCancellation"
                    form_def.encounterTypeName = base_name
                continue

            # 9. Default: Encounter (general encounter)
            form_def.formType = "Encounter"
            form_def.encounterTypeName = form_def.name.strip()

    def _find_matching_program(
        self,
        base: str,
        program_names: dict[str, str],
    ) -> str | None:
        """Find a program name that matches a base string.

        Uses fuzzy substring matching.
        """
        base_lower = base.lower().strip()

        # Exact match
        if base_lower in program_names:
            return program_names[base_lower]

        # Check if base is a prefix of any program
        for pname_lower, pname in program_names.items():
            if pname_lower.startswith(base_lower) or base_lower.startswith(pname_lower):
                return pname

        # Check if any word from base appears in program names
        base_words = set(base_lower.split())
        best_match: str | None = None
        best_score = 0
        for pname_lower, pname in program_names.items():
            pwords = set(pname_lower.split())
            overlap = len(base_words & pwords)
            if overlap > best_score:
                best_score = overlap
                best_match = pname

        if best_score > 0:
            return best_match

        # Last resort: first program
        if program_names:
            return next(iter(program_names.values()))

        return None

    def _resolve_groups(self) -> list[str]:
        """Determine user groups from all sources."""
        groups: list[str] = []
        seen: set[str] = set()

        # Always include "Everyone"
        groups.append("Everyone")
        seen.add("everyone")

        # From user sheet
        for g in self._user_groups:
            gl = g.lower()
            if gl not in seen:
                groups.append(g)
                seen.add(gl)

        # From permissions sheet
        for g in self._permission_groups:
            gl = g.lower()
            if gl not in seen:
                groups.append(g)
                seen.add(gl)

        return groups

    def _resolve_address_levels(self) -> list[dict[str, Any]]:
        """Determine address level types."""
        if self._location_hierarchy:
            return self._location_hierarchy

        # Try to extract from summary
        hierarchy_text = self._summary.get("location_hierarchy", "")
        if hierarchy_text:
            return self._parse_hierarchy_text(hierarchy_text)

        return []

    def _parse_hierarchy_text(self, text: str) -> list[dict[str, Any]]:
        """Parse a location hierarchy from text like 'Block -> Village -> Para',
        'State: Karnataka\\nDistrict: Bengaluru', or 'State,\\nDistrict,\\nBlock'.
        """
        levels: list[dict[str, Any]] = []

        # Try arrow-separated: "Block -> Village -> Para"
        if "->" in text:
            parts = [p.strip() for p in text.split("->") if p.strip()]
            total = len(parts)
            for i, part in enumerate(parts):
                level_entry: dict[str, Any] = {
                    "name": part,
                    "level": total - i,
                }
                if i > 0:
                    level_entry["parent"] = parts[i - 1]
                levels.append(level_entry)
            return levels

        # Try colon-separated lines: "State: Karnataka\nDistrict: Bengaluru"
        lines = text.split("\n")
        has_colons = any(":" in line for line in lines if line.strip())
        if has_colons:
            for line in lines:
                line = line.strip().lstrip("\u2022\t -")
                if ":" in line:
                    name = line.split(":")[0].strip()
                    name = re.sub(r"\s*/.*$", "", name).strip()  # Remove "/ Region" etc.
                    if name:
                        levels.append({"name": name})
        else:
            # Try comma and/or newline separated: "State,\nDistrict,\nBlock"
            # First rejoin and split by comma (handles "State,\nDistrict" format)
            joined = text.replace("\n", ",")
            parts = [p.strip().rstrip(",") for p in joined.split(",")]
            for part in parts:
                part = part.strip().lstrip("\u2022\t -")
                if part:
                    levels.append({"name": part})

        total = len(levels)
        for i, level_entry in enumerate(levels):
            level_entry["level"] = total - i
            if i > 0:
                level_entry["parent"] = levels[i - 1]["name"]

        return levels

    def _resolve_program_encounter_mappings(self) -> list[dict[str, Any]]:
        """Build program-encounter type mappings."""
        pe_map: dict[str, set[str]] = {}

        # From parsed forms
        for form_def in self._forms:
            if form_def.formType in ("ProgramEncounter", "ProgramEncounterCancellation"):
                if form_def.programName and form_def.encounterTypeName:
                    pe_map.setdefault(form_def.programName, set()).add(
                        form_def.encounterTypeName
                    )

        # From program encounters meta
        for pe in self._program_encounters_meta:
            prog = pe.get("program_name", "").strip()
            enc = pe["encounter_name"].strip()
            if prog:
                pe_map.setdefault(prog, set()).add(enc)

        return [
            {"program": prog, "encounterTypes": sorted(ets)}
            for prog, ets in pe_map.items()
        ]

    def _resolve_general_encounter_types(self) -> list[str]:
        """Find encounter types not linked to any program."""
        general: list[str] = []
        for form_def in self._forms:
            if form_def.formType in ("Encounter", "IndividualEncounterCancellation"):
                if form_def.encounterTypeName:
                    if form_def.encounterTypeName not in general:
                        general.append(form_def.encounterTypeName)
        return general


# ---------------------------------------------------------------------------
# Convenience function
# ---------------------------------------------------------------------------


def parse_srs_excel(file_path: str | Path, sector: str | None = None) -> SRSData:
    """Parse an SRS Excel file and return structured SRSData.

    This is the primary entry point for use by other modules.
    Auto-detects canonical template format and uses deterministic parser if matched.

    Args:
        file_path: Path to the SRS Excel file.
        sector: Optional sector hint (MCH, Education, etc.) for better classification.
    """
    # Auto-detect canonical template — deterministic, zero LLM
    from app.services.canonical_srs_parser import is_canonical_template, parse_canonical_srs
    if is_canonical_template(str(file_path)):
        logger.info("Detected canonical SRS template: %s", file_path)
        srs_data, errors = parse_canonical_srs(str(file_path))
        if errors:
            logger.warning(
                "Canonical SRS parsed with %d validation error(s): %s",
                len(errors), "; ".join(errors[:5]),
            )
        return srs_data

    # Fallback: heuristic parser for free-form SRS Excel
    parser = SRSParser(file_path)
    parser._sector = sector
    return parser.parse()


def parse_multiple_srs_excels(file_paths: list[str | Path], sector: str | None = None) -> SRSData:
    """Parse multiple SRS Excel files and merge them into one SRSData.

    Useful when scoping doc and modelling doc are separate files.
    The first file is treated as primary; subsequent files enrich it.
    """
    if not file_paths:
        raise ValueError("At least one file path is required")

    if len(file_paths) == 1:
        return parse_srs_excel(file_paths[0], sector=sector)

    # Parse primary file
    primary = SRSParser(file_paths[0])
    primary._sector = sector
    primary.parse()  # fills internal state

    # Parse additional files and merge their data into primary
    for fp in file_paths[1:]:
        secondary = SRSParser(fp)
        secondary._sector = sector
        secondary.parse()

        # Merge subject types (avoid duplicates by name)
        existing_st = {st["name"].lower() for st in (primary._modelling.get("subject_types") or [])}
        for st in (secondary._modelling.get("subject_types") or []):
            if st["name"].lower() not in existing_st:
                if not primary._modelling.get("subject_types"):
                    primary._modelling["subject_types"] = []
                primary._modelling["subject_types"].append(st)
                existing_st.add(st["name"].lower())

        # Merge programs meta
        existing_progs = {pm["name"].lower() for pm in primary._programs_meta}
        for pm in secondary._programs_meta:
            if pm["name"].lower() not in existing_progs:
                primary._programs_meta.append(pm)
                existing_progs.add(pm["name"].lower())

        # Merge encounters meta
        existing_enc = {em["name"].lower() for em in primary._encounters_meta}
        for em in secondary._encounters_meta:
            if em["name"].lower() not in existing_enc:
                primary._encounters_meta.append(em)
                existing_enc.add(em["name"].lower())

        # Merge program encounters meta
        for pem in secondary._program_encounters_meta:
            primary._program_encounters_meta.append(pem)

        # Merge location hierarchy if primary doesn't have one
        if not primary._location_hierarchy and secondary._location_hierarchy:
            primary._location_hierarchy = secondary._location_hierarchy

        # Merge visit schedules
        if secondary._visit_schedules:
            primary._visit_schedules.extend(secondary._visit_schedules)

        # Merge forms (from additional scoping sheets)
        existing_forms = {f.name.lower() for f in primary._forms}
        for f in secondary._forms:
            if f.name.lower() not in existing_forms:
                primary._forms.append(f)
                existing_forms.add(f.name.lower())

    # Rebuild with merged data
    return primary._build_srs_data()
