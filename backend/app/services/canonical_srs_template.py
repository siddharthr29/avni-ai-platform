"""Generate and export the canonical SRS template for 100% correct bundle generation.

The canonical template has a fixed structure that maps 1:1 to Avni's data model:
  - Sheet 1: Modelling (subject types, programs, encounter types, forms)
  - Sheet 2: Location Hierarchy
  - Sheet 3: Groups
  - Sheet 4: Visit Schedules
  - Sheets 5+: One sheet per form (named exactly as Form Name in Modelling)

Users fill the template, upload it, and get a deterministic, LLM-free bundle.
"""

from __future__ import annotations

import io
import csv
import logging
from typing import Any

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.worksheet import Worksheet

from app.models.schemas import SRSData

logger = logging.getLogger(__name__)

# ---------------------------------------------------------------------------
# Constants
# ---------------------------------------------------------------------------

VALID_FORM_TYPES = [
    "IndividualProfile", "ProgramEnrolment", "ProgramExit",
    "ProgramEncounter", "ProgramEncounterCancellation",
    "Encounter", "IndividualEncounterCancellation",
]

VALID_DATA_TYPES = [
    "Text", "Numeric", "Coded", "Date", "DateTime", "Time",
    "Notes", "Image", "PhoneNumber", "Location", "Subject",
    "QuestionGroup", "Duration", "Id", "Audio", "Video", "File",
]

VALID_ENTITY_TYPES = ["SubjectType", "Program", "EncounterType"]

VALID_SUBJECT_KINDS = ["Person", "Household", "Group"]

# Styling
HEADER_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
HEADER_FONT = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
REQUIRED_FILL = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
EXAMPLE_FONT = Font(name="Calibri", italic=True, color="808080", size=10)
THIN_BORDER = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin"),
)

# ---------------------------------------------------------------------------
# Modelling sheet columns
# ---------------------------------------------------------------------------

MODELLING_HEADERS = [
    "Entity Type",          # SubjectType / Program / EncounterType
    "Name",                 # e.g. Individual, Maternal Health, ANC Visit
    "Parent (Type/Program)",# For SubjectType: Person/Household/Group. For EncounterType: program name or blank
    "Form Name",            # Registration, ANC Enrolment, ANC Visit
    "Form Type",            # IndividualProfile, ProgramEnrolment, ProgramEncounter, Encounter, etc.
    "Colour",               # hex colour for programs, optional
]

# Examples use generic names — delete these and replace with your org's entities
MODELLING_EXAMPLES = [
    ["SubjectType", "Individual", "Person", "Registration", "IndividualProfile", ""],
    ["Program", "My Program", "", "My Program Enrolment", "ProgramEnrolment", "#E91E63"],
    ["EncounterType", "Follow-up Visit", "My Program", "Follow-up Visit", "ProgramEncounter", ""],
    ["EncounterType", "General Check-up", "", "General Check-up", "Encounter", ""],
]

# ---------------------------------------------------------------------------
# Location Hierarchy sheet columns
# ---------------------------------------------------------------------------

LOCATION_HEADERS = ["Level Name", "Level Number", "Parent Level"]
LOCATION_EXAMPLES = [
    ["State", "3", ""],
    ["District", "2", "State"],
    ["Block", "1", "District"],
]

# ---------------------------------------------------------------------------
# Groups sheet columns
# ---------------------------------------------------------------------------

GROUPS_HEADERS = ["Group Name", "Has All Privileges"]
GROUPS_EXAMPLES = [["Everyone", "Yes"]]

# ---------------------------------------------------------------------------
# Visit Schedules sheet columns
# ---------------------------------------------------------------------------

VISIT_HEADERS = [
    "After Form",           # Which form triggers this schedule
    "Schedule Encounter",   # Which encounter type to schedule
    "Visit Name",           # Display name for the visit (optional)
    "Due Days",             # Days after trigger until due
    "Overdue Days",         # Days after trigger until overdue
    "On Cancellation",      # Reschedule / Close (optional)
]

# Example: schedule a follow-up 28 days after each visit, overdue at 35 days
VISIT_EXAMPLES = [
    ["Follow-up Visit", "Follow-up Visit", "Follow-up Visit 2", "28", "35", "Reschedule"],
]

# ---------------------------------------------------------------------------
# Form sheet columns
# ---------------------------------------------------------------------------

FORM_HEADERS = [
    "Page/Section",     # Maps to formElementGroup name
    "Field Name",       # Concept name
    "Data Type",        # Numeric, Text, Coded, Date, etc.
    "Mandatory",        # Yes/No
    "Options",          # Semicolon-separated for Coded fields
    "Selection Type",   # Single/Multi for Coded fields
    "Unit",             # For Numeric fields
    "Min Value",        # For Numeric fields
    "Max Value",        # For Numeric fields
    "Show When",        # Structured skip logic
    "Hide When",        # Structured skip logic
    "QG Parent",        # QuestionGroup parent field name
]

# ---------------------------------------------------------------------------
# Decisions sheet columns
# ---------------------------------------------------------------------------

DECISIONS_HEADERS = [
    "Form Name",            # Which form this decision runs on
    "When",                 # Condition: same format as Show When (or "ALWAYS")
    "Set Field",            # Target concept name (decision output)
    "To Value",             # Value to set — can reference fields with {FieldName}
    "Scope",                # encounter / enrolment / registration
]

DECISIONS_EXAMPLES = [
    # Simple condition → set value
    ["Follow-up Visit", "Status = Referred", "Referral Flag", "Yes", "encounter"],
    # Always run — auto-stamp metadata
    ["Registration", "ALWAYS", "Registration Date", "{CURRENT_DATE}", "registration"],
    # Formula
    ["Follow-up Visit", "ALWAYS", "BMI", "{Weight} / ({Height} / 100) ^ 2", "encounter"],
    # Classification based on ranges
    ["Follow-up Visit", "Score < 50", "Risk Category", "High", "encounter"],
    ["Follow-up Visit", "Score >= 50 AND Score < 80", "Risk Category", "Medium", "encounter"],
    ["Follow-up Visit", "Score >= 80", "Risk Category", "Low", "encounter"],
    # Complications builder: check if value matches → add to list
    ["Follow-up Visit", "Systolic >= 140", "Complications", "Hypertension", "encounter"],
    ["Follow-up Visit", "Temperature > 38", "Complications", "Fever", "encounter"],
    # Copy field with formatting
    ["Registration", "ALWAYS", "Name Display", "{Name} ({Gender})", "registration"],
]

# ---------------------------------------------------------------------------
# Eligibility sheet columns
# ---------------------------------------------------------------------------

ELIGIBILITY_HEADERS = [
    "Program",              # Which program this eligibility rule applies to
    "Condition",            # Structured condition using same format as Show When
]

ELIGIBILITY_EXAMPLES = [
    # Simple age + gender check
    ["My Program", "Gender = Female AND Age > 15"],
    # Age-only check
    ["My Program", "Age < 5"],
]

# ---------------------------------------------------------------------------
# Report Cards sheet columns
# ---------------------------------------------------------------------------

REPORT_CARDS_HEADERS = [
    "Card Name",            # Display name on dashboard
    "Card Type",            # Standard type OR "Custom"
    "Description",          # Short description (optional)
    "Subject Type",         # Which subject type to filter (for standard cards)
    "Program",              # Filter by program (optional)
    "Encounter Type",       # Filter by encounter type (optional)
    "Recent Duration",      # For Recent cards: "1 day", "1 week", "1 month"
    "Filter Condition",     # For Custom cards: structured filter expression
    "Colour",               # Hex colour (optional)
    "Nested",               # Yes/No — whether card contains sub-cards
]

VALID_CARD_TYPES = [
    "Total", "ScheduledVisits", "OverdueVisits",
    "RecentRegistrations", "RecentEnrolments", "RecentVisits",
    "DueChecklist", "Tasks", "CallTasks", "OpenSubjectTasks",
    "PendingApproval", "Approved", "Rejected", "Comments",
    "Custom",
]

REPORT_CARDS_EXAMPLES = [
    # Standard cards
    ["Total Beneficiaries", "Total", "", "Individual", "", "", "", "", "#4CAF50", "No"],
    ["Scheduled Visits", "ScheduledVisits", "Visits due this week", "Individual", "My Program", "Follow-up Visit", "", "", "#FF9800", "No"],
    ["Overdue Visits", "OverdueVisits", "Visits past due date", "Individual", "My Program", "Follow-up Visit", "", "", "#F44336", "No"],
    ["Recent Registrations", "RecentRegistrations", "Registered this week", "Individual", "", "", "1 week", "", "#2196F3", "No"],
    ["Recent Enrolments", "RecentEnrolments", "Enrolled this month", "Individual", "My Program", "", "1 month", "", "#9C27B0", "No"],
    # Custom card with filter condition
    ["High Risk Cases", "Custom", "Cases with High Risk status", "Individual", "My Program", "", "", "Risk Category = High", "#E91E63", "No"],
    ["Active Female Beneficiaries", "Custom", "", "Individual", "", "", "", "Gender = Female AND Status = Active", "#00BCD4", "No"],
]


# ---------------------------------------------------------------------------
# Template generation
# ---------------------------------------------------------------------------


def _style_header_row(ws: Worksheet, headers: list[str]) -> None:
    """Apply header styling to the first row."""
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = THIN_BORDER
    ws.row_dimensions[1].height = 30


def _add_example_rows(ws: Worksheet, examples: list[list[str]], start_row: int = 2) -> None:
    """Add example rows with italic grey font."""
    for row_idx, row_data in enumerate(examples, start_row):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.font = EXAMPLE_FONT
            cell.border = THIN_BORDER


def _auto_width(ws: Worksheet, headers: list[str]) -> None:
    """Auto-size columns based on header length (minimum 12 chars)."""
    for col_idx, header in enumerate(headers, 1):
        width = max(len(header) + 4, 14)
        ws.column_dimensions[get_column_letter(col_idx)].width = width


def _add_data_validation(ws: Worksheet, col_idx: int, values: list[str], max_row: int = 200) -> None:
    """Add dropdown data validation to a column."""
    formula = '"' + ",".join(values) + '"'
    dv = DataValidation(type="list", formula1=formula, allow_blank=True)
    dv.error = "Please select a valid value"
    dv.prompt = "Select from dropdown"
    col_letter = get_column_letter(col_idx)
    dv.add(f"{col_letter}2:{col_letter}{max_row}")
    ws.add_data_validation(dv)


def _build_instructions_sheet(wb: openpyxl.Workbook) -> None:
    """Add an Instructions sheet with comprehensive usage guide."""
    ws = wb.create_sheet("Instructions", 0)

    TITLE_FONT = Font(name="Calibri", bold=True, size=16, color="1F4E79")
    HEADING_FONT = Font(name="Calibri", bold=True, size=12, color="1F4E79")
    SUBHEAD_FONT = Font(name="Calibri", bold=True, size=11)
    BODY_FONT = Font(name="Calibri", size=10)
    CODE_FONT = Font(name="Consolas", size=10)
    HINT_FONT = Font(name="Calibri", italic=True, color="666666", size=10)
    WARN_FONT = Font(name="Calibri", bold=True, color="CC0000", size=10)

    # (col_a_text, col_b_text, col_a_font_override)
    rows: list[tuple[str, str, Font | None]] = [
        ("Avni SRS Canonical Template — Instructions", "", TITLE_FONT),
        ("", "", None),
        ("OVERVIEW", "", HEADING_FONT),
        ("This template maps 1:1 to Avni's data model. Fill it in and upload to", "", None),
        ("generate a 100% correct implementation bundle — with rules, skip logic,", "", None),
        ("visit schedules, and validations. No AI guessing. Works for orgs with 1", "", None),
        ("form or 50+ forms.", "", None),
        ("", "", None),

        # ── STEP 1 ──
        ("STEP 1: MODELLING SHEET (required)", "", HEADING_FONT),
        ("Define ALL subject types, programs, encounter types, and their forms.", "", None),
        ("Each row = one entity and the form that goes with it.", "", None),
        ("", "", None),
        ("Column guide:", "", SUBHEAD_FONT),
        ("  Entity Type", "SubjectType / Program / EncounterType", CODE_FONT),
        ("  Name", "The entity name (e.g. Individual, Maternal Health, ANC Visit)", CODE_FONT),
        ("  Parent (Type/Program)", "SubjectType → Person, Household, or Group", CODE_FONT),
        ("", "EncounterType → program name it belongs to (blank = standalone encounter)", HINT_FONT),
        ("  Form Name", "Name of the form (must match a sheet name in this workbook)", CODE_FONT),
        ("  Form Type", "IndividualProfile, ProgramEnrolment, ProgramEncounter, Encounter, etc.", CODE_FONT),
        ("  Colour", "Hex colour for programs, e.g. #E91E63 (optional)", CODE_FONT),
        ("", "", None),
        ("Form Type quick reference:", "", SUBHEAD_FONT),
        ("  IndividualProfile", "Registration form for a subject type", CODE_FONT),
        ("  ProgramEnrolment", "Enrolment form when subject enters a program", CODE_FONT),
        ("  ProgramEncounter", "Visit/encounter form within a program", CODE_FONT),
        ("  Encounter", "Standalone encounter (not under any program)", CODE_FONT),
        ("  ProgramExit", "Exit form when subject leaves a program (auto-generated if missing)", CODE_FONT),
        ("  ProgramEncounterCancellation", "Cancellation form for program encounters (auto-generated if missing)", CODE_FONT),
        ("  IndividualEncounterCancellation", "Cancellation form for standalone encounters (auto-generated if missing)", CODE_FONT),
        ("", "", None),
        ("Example A — Health sector (1 subject, 1 program, 2 encounters):", "", SUBHEAD_FONT),
        ("  SubjectType | Individual | Person | Registration | IndividualProfile", "", CODE_FONT),
        ("  Program | Maternal Health | | MH Enrolment | ProgramEnrolment | #E91E63", "", CODE_FONT),
        ("  EncounterType | ANC Visit | Maternal Health | ANC Visit | ProgramEncounter", "", CODE_FONT),
        ("  EncounterType | Home Visit | | Home Visit | Encounter", "", CODE_FONT),
        ("  → 4 rows, 4 form sheets needed", "", HINT_FONT),
        ("", "", None),
        ("Example B — Education sector (2 subjects, 1 program, 3 encounters):", "", SUBHEAD_FONT),
        ("  SubjectType | Student | Person | Student Registration | IndividualProfile", "", CODE_FONT),
        ("  SubjectType | School | Group | School Registration | IndividualProfile", "", CODE_FONT),
        ("  Program | Remedial Learning | | RL Enrolment | ProgramEnrolment | #2196F3", "", CODE_FONT),
        ("  EncounterType | Baseline Assessment | Remedial Learning | Baseline Assessment | ProgramEncounter", "", CODE_FONT),
        ("  EncounterType | Daily Attendance | Remedial Learning | Daily Attendance | ProgramEncounter", "", CODE_FONT),
        ("  EncounterType | School Inspection | | School Inspection | Encounter", "", CODE_FONT),
        ("  → 6 rows, 6 form sheets needed", "", HINT_FONT),
        ("", "", None),
        ("Example C — Livelihoods / WASH / complex org (3 subjects, 2 programs, 8 encounters):", "", SUBHEAD_FONT),
        ("  SubjectType | Individual | Person | Individual Registration | IndividualProfile", "", CODE_FONT),
        ("  SubjectType | Household | Household | Household Registration | IndividualProfile", "", CODE_FONT),
        ("  SubjectType | Village | Group | Village Registration | IndividualProfile", "", CODE_FONT),
        ("  Program | Nutrition | | Nutrition Enrolment | ProgramEnrolment | #4CAF50", "", CODE_FONT),
        ("  Program | Livelihoods | | Livelihoods Enrolment | ProgramEnrolment | #FF9800", "", CODE_FONT),
        ("  EncounterType | Growth Monitoring | Nutrition | Growth Monitoring | ProgramEncounter", "", CODE_FONT),
        ("  EncounterType | MUAC Assessment | Nutrition | MUAC Assessment | ProgramEncounter", "", CODE_FONT),
        ("  EncounterType | Supplementary Food | Nutrition | Supplementary Food | ProgramEncounter", "", CODE_FONT),
        ("  EncounterType | Skills Training | Livelihoods | Skills Training | ProgramEncounter", "", CODE_FONT),
        ("  EncounterType | Micro-enterprise Review | Livelihoods | Micro-enterprise Review | ProgramEncounter", "", CODE_FONT),
        ("  EncounterType | Household Survey | | Household Survey | Encounter", "", CODE_FONT),
        ("  EncounterType | Spot Check | | Spot Check | Encounter", "", CODE_FONT),
        ("  EncounterType | Village Meeting | | Village Meeting | Encounter", "", CODE_FONT),
        ("  → 13 rows, 13 form sheets needed (orgs with 30-50+ forms just keep adding rows)", "", HINT_FONT),
        ("", "", None),

        # ── STEP 2 ──
        ("STEP 2: CREATE FORM SHEETS", "", HEADING_FONT),
        ("For EACH Form Name in the Modelling sheet, create a new sheet in this workbook.", "", None),
        ("The sheet name must EXACTLY match the Form Name.", "", None),
        ("", "", None),
        ("IMPORTANT: Excel limits sheet names to 31 characters.", WARN_FONT and "", WARN_FONT),
        ("If your form name is longer, truncate it to 31 chars in both the Modelling", "", None),
        ("sheet AND the sheet tab name.", "", None),
        ("", "", None),
        ("How to add a new form sheet:", "", SUBHEAD_FONT),
        ("  1. Right-click any sheet tab at the bottom → 'Insert Sheet'", "", None),
        ("  2. Rename it to EXACTLY match the Form Name from the Modelling sheet", "", None),
        ("  3. Copy the header row from the 'Example Form' sheet", "", None),
        ("  4. Fill in your fields (one row per field)", "", None),
        ("", "", None),
        ("For orgs with many forms (20, 30, 50+), just keep adding sheets.", "", HINT_FONT),
        ("There is no limit on the number of form sheets.", "", HINT_FONT),
        ("", "", None),
        ("Form sheet columns:", "", SUBHEAD_FONT),
        ("  Page/Section", "Groups fields into pages/tabs on the mobile form (required)", CODE_FONT),
        ("  Field Name", "The concept name — must be unique across the entire org (required)", CODE_FONT),
        ("  Data Type", "Text, Numeric, Coded, Date, Notes, PhoneNumber, Image, Location, Subject, QuestionGroup, etc. (required)", CODE_FONT),
        ("  Mandatory", "Yes or No (required)", CODE_FONT),
        ("  Options", "For Coded fields only — semicolon-separated values, e.g. Yes; No; Unknown", CODE_FONT),
        ("  Selection Type", "Single (radio buttons) or Multi (checkboxes) — for Coded fields", CODE_FONT),
        ("  Unit", "For Numeric fields — e.g. kg, cm, mmHg", CODE_FONT),
        ("  Min Value", "For Numeric fields — minimum allowed value (generates validation rule)", CODE_FONT),
        ("  Max Value", "For Numeric fields — maximum allowed value (generates validation rule)", CODE_FONT),
        ("  Show When", "Skip logic: show this field when condition is true (see format below)", CODE_FONT),
        ("  Hide When", "Skip logic: hide this field when condition is true", CODE_FONT),
        ("  QG Parent", "For fields inside a QuestionGroup — name of the parent QG field", CODE_FONT),
        ("", "", None),

        # ── STEP 3 ──
        ("STEP 3: LOCATION HIERARCHY (optional)", "", HEADING_FONT),
        ("Define address levels for the org. Defaults to State > District > Block if empty.", "", None),
        ("Higher Level Number = higher in hierarchy (e.g. State=3, District=2, Block=1).", "", None),
        ("", "", None),

        # ── STEP 4 ──
        ("STEP 4: GROUPS (optional)", "", HEADING_FONT),
        ("Define user groups. 'Everyone' is always included even if you don't add it.", "", None),
        ("Set 'Has All Privileges' to Yes for groups that should have full access.", "", None),
        ("", "", None),

        # ── STEP 5 ──
        ("STEP 5: VISIT SCHEDULES (optional)", "", HEADING_FONT),
        ("Define recurring visit scheduling. Each row says:", "", None),
        ("'After completing [After Form], schedule [Schedule Encounter] in [Due Days] days.'", "", None),
        ("", "", None),
        ("Example:", "", SUBHEAD_FONT),
        ("  After Form: Follow-up Visit | Schedule: Follow-up Visit | Due: 28 | Overdue: 35", "Next visit 28 days after current one", CODE_FONT),
        ("  After Form: My Program Enrolment | Schedule: Follow-up Visit | Due: 0 | Overdue: 7", "First visit immediately after enrolment", CODE_FONT),
        ("  After Form: Monthly Check | Schedule: Monthly Check | Due: 30 | Overdue: 37", "Monthly recurring visits", CODE_FONT),
        ("  On Cancellation = Reschedule means a cancelled visit auto-reschedules.", "", HINT_FONT),
        ("", "", None),

        # ── STEP 6 ──
        ("STEP 6: DECISIONS (optional)", "", HEADING_FONT),
        ("Auto-set field values based on conditions. Each row says:", "", None),
        ("'When [condition] is true on [form], set [field] to [value].'", "", None),
        ("", "", None),
        ("Column guide:", "", SUBHEAD_FONT),
        ("  Form Name", "Must match a form in the Modelling sheet", CODE_FONT),
        ("  When", "Condition (same format as Show When), or ALWAYS to run unconditionally", CODE_FONT),
        ("  Set Field", "The concept name to write the decision value into", CODE_FONT),
        ("  To Value", "Plain text, or {FieldName} to reference another field's value", CODE_FONT),
        ("", "{Weight} / ({Height} / 100) ^ 2  — arithmetic with field references", HINT_FONT),
        ("", "{CURRENT_DATE}  — today's date", HINT_FONT),
        ("", "{CURRENT_USER}  — logged-in user's name", HINT_FONT),
        ("  Scope", "encounter / enrolment / registration — where to store the decision", CODE_FONT),
        ("", "", None),
        ("Use cases:", "", SUBHEAD_FONT),
        ("  Auto-classify: Score < 50 → Risk Category = High", "", None),
        ("  Auto-calculate: BMI = {Weight} / ({Height} / 100) ^ 2", "", None),
        ("  Complications: Systolic >= 140 → Complications = Hypertension", "Multiple rows with same Set Field build a list", None),
        ("  Auto-stamp: ALWAYS → Registration Date = {CURRENT_DATE}", "", None),
        ("  Copy field: ALWAYS → Display Name = {Name} ({Gender})", "", None),
        ("", "", None),

        # ── STEP 7 ──
        ("STEP 7: ELIGIBILITY (optional)", "", HEADING_FONT),
        ("Define who can enrol in each program. Each row says:", "", None),
        ("'Only allow enrolment in [program] when [condition] is true.'", "", None),
        ("", "", None),
        ("The condition uses the same format as Show When, plus special keywords:", "", None),
        ("  Gender = Female", "Only females can enrol", CODE_FONT),
        ("  Age > 15", "Only above 15 years", CODE_FONT),
        ("  Gender = Female AND Age > 15", "Compound condition", CODE_FONT),
        ("If no eligibility row for a program, anyone can enrol.", "", HINT_FONT),
        ("", "", None),

        # ── STEP 8 ──
        ("STEP 8: REPORT CARDS (optional)", "", HEADING_FONT),
        ("Define dashboard cards. Two kinds:", "", None),
        ("  Standard cards: built-in types (Total, ScheduledVisits, OverdueVisits, etc.)", "", None),
        ("  Custom cards: your own filter condition to count matching subjects", "", None),
        ("", "", None),
        ("Standard card types:", "", SUBHEAD_FONT),
        ("  Total", "Count of all registered subjects of a type", CODE_FONT),
        ("  ScheduledVisits", "Visits due (not yet done)", CODE_FONT),
        ("  OverdueVisits", "Visits past their due date", CODE_FONT),
        ("  RecentRegistrations", "Subjects registered within Recent Duration", CODE_FONT),
        ("  RecentEnrolments", "Subjects enrolled within Recent Duration", CODE_FONT),
        ("  RecentVisits", "Visits completed within Recent Duration", CODE_FONT),
        ("  DueChecklist", "Checklist items due", CODE_FONT),
        ("  Tasks / CallTasks / OpenSubjectTasks", "Task-related cards", CODE_FONT),
        ("  PendingApproval / Approved / Rejected", "Approval workflow cards", CODE_FONT),
        ("", "", None),
        ("Custom card filter (same format as skip logic):", "", SUBHEAD_FONT),
        ("  Risk Category = High", "Count subjects where Risk Category is High", CODE_FONT),
        ("  Gender = Female AND Status = Active", "Compound filter", CODE_FONT),
        ("  Status IN (Active, Pending)", "Multiple values", CODE_FONT),
        ("", "", None),

        # ── SKIP LOGIC ──
        ("SKIP LOGIC FORMAT (Show When / Hide When / Decision When / Filter Condition)", "", HEADING_FONT),
        ("Write conditions in this structured format — they are parsed deterministically,", "", None),
        ("no AI guessing. Generated rules work on both Avni Android app and webapp.", "", None),
        ("", "", None),
        ("  FieldName = Value", "Show when FieldName equals Value (for Coded fields)", CODE_FONT),
        ("  FieldName != Value", "Show when FieldName does NOT equal Value", CODE_FONT),
        ("  FieldName > 5", "Show when FieldName is greater than 5", CODE_FONT),
        ("  FieldName < 10", "Show when FieldName is less than 10", CODE_FONT),
        ("  FieldName >= 18", "Show when FieldName is greater than or equal to 18", CODE_FONT),
        ("  FieldName <= 100", "Show when FieldName is less than or equal to 100", CODE_FONT),
        ("  FieldName IS EMPTY", "Show when FieldName has no value", CODE_FONT),
        ("  FieldName IS NOT EMPTY", "Show when FieldName has a value", CODE_FONT),
        ("  FieldName CONTAINS Value", "Show when multi-select FieldName contains Value", CODE_FONT),
        ("  FieldName IN (A, B, C)", "Show when FieldName is one of A, B, or C", CODE_FONT),
        ("  Cond1 AND Cond2", "Both conditions must be true", CODE_FONT),
        ("  Cond1 OR Cond2", "Either condition must be true", CODE_FONT),
        ("", "", None),
        ("Example skip logic rows in a form sheet:", "", SUBHEAD_FONT),
        ("  Field: 'Other Reason'   → Show When: Reason = Other", "Shows 'Other' text box when Other is selected", CODE_FONT),
        ("  Field: 'Follow-up Notes' → Show When: Outcome = Needs Follow-up", "Shows notes only when follow-up is needed", CODE_FONT),
        ("  Field: 'Spouse Name'    → Show When: Marital Status = Married", "Shows only for married beneficiaries", CODE_FONT),
        ("  Field: 'Score Summary'  → Show When: Score IS NOT EMPTY", "Shows only after score is entered", CODE_FONT),
        ("  Field: 'Details'        → Show When: Age >= 18 AND Status = Active", "Compound: both must be true", CODE_FONT),
        ("", "", None),
        ("IMPORTANT: The FieldName in skip logic must match a Field Name", "", WARN_FONT),
        ("in the SAME form sheet. Cross-form references are not supported.", "", WARN_FONT),
        ("", "", None),

        # ── DATA TYPES ──
        ("DATA TYPE REFERENCE", "", HEADING_FONT),
        ("  Text", "Short text input", CODE_FONT),
        ("  Numeric", "Number input — add Unit, Min Value, Max Value for validation", CODE_FONT),
        ("  Coded", "Dropdown / radio / checkbox — requires Options and Selection Type", CODE_FONT),
        ("  Date", "Date picker", CODE_FONT),
        ("  DateTime", "Date + time picker", CODE_FONT),
        ("  Time", "Time picker", CODE_FONT),
        ("  Notes", "Long text / multi-line input", CODE_FONT),
        ("  PhoneNumber", "Phone keypad input on mobile", CODE_FONT),
        ("  Image", "Camera / image picker on mobile", CODE_FONT),
        ("  Video", "Video capture on mobile", CODE_FONT),
        ("  Audio", "Audio recording on mobile", CODE_FONT),
        ("  Location", "GPS / location hierarchy picker", CODE_FONT),
        ("  Subject", "Link to another registered subject (renders subject picker)", CODE_FONT),
        ("  QuestionGroup", "Repeating group of fields — child fields use QG Parent column", CODE_FONT),
        ("  Id", "Auto-generated identifier", CODE_FONT),
        ("  Duration", "Duration input (days/months/years)", CODE_FONT),
        ("  File", "File attachment", CODE_FONT),
        ("", "", None),

        # ── WHAT'S AUTO-GENERATED ──
        ("WHAT GETS AUTO-GENERATED", "", HEADING_FONT),
        ("You only need to define the primary forms. The system auto-generates:", "", None),
        ("  - Cancellation forms for every encounter type (with Cancel Reason field)", "", None),
        ("  - Exit forms for every program (with Exit Reason and Exit Date fields)", "", None),
        ("  - Skip logic rules (JS + declarative) from Show When / Hide When columns", "", None),
        ("  - Visit schedule rules (JS + declarative) from the Visit Schedules sheet", "", None),
        ("  - Decision rules (JS) from the Decisions sheet — auto-calc, classify, stamp", "", None),
        ("  - Eligibility rules (JS) from the Eligibility sheet — who can enrol", "", None),
        ("  - Report cards + dashboard from the Report Cards sheet", "", None),
        ("  - Numeric validation rules from Min Value / Max Value columns", "", None),
        ("  - Date validation rules for encounter forms", "", None),
        ("  - Group privileges for all groups × subject types × programs × encounter types", "", None),
        ("  - Concepts with stable UUIDs (same field name always gets same UUID)", "", None),
        ("  - Form mappings linking forms to subject types, programs, and encounter types", "", None),
        ("", "", None),

        # ── COMMON MISTAKES ──
        ("COMMON MISTAKES TO AVOID", "", HEADING_FONT),
        ("  1. Form sheet name doesn't match Form Name in Modelling sheet", "Parser will report 'Missing form sheet' error", None),
        ("  2. Same Field Name used twice in one form", "Server rejects duplicate concepts in a form", None),
        ("  3. Coded field without Options", "Coded fields must have at least one option", None),
        ("  4. EncounterType without a parent program listed as ProgramEncounter", "Standalone encounters should use Form Type = Encounter", None),
        ("  5. Leaving example rows in when uploading", "Delete the grey italic rows first", None),
        ("  6. Same concept name with different Data Types in different forms", "A field named 'Weight' must be Numeric everywhere", None),
        ("  7. Skip logic referencing a field in a different form", "Only same-form references are supported in this template", None),
        ("  8. Sheet name > 31 characters", "Excel truncates — match exactly in both places", None),
        ("", "", None),

        # ── UPLOAD ──
        ("UPLOADING", "", HEADING_FONT),
        ("When ready, upload this .xlsx file to Avni AI platform:", "", None),
        ("  - Drag and drop onto the SRS Builder page, or", "", None),
        ("  - Use the 'Upload SRS Excel' option in chat, or", "", None),
        ("  - POST to /api/bundle/generate-from-excel", "", None),
        ("The system auto-detects this as a canonical template and uses the", "", None),
        ("deterministic parser (zero LLM, zero guessing).", "", None),
    ]

    for row_idx, (col_a, col_b, font_override) in enumerate(rows, 1):
        cell_a = ws.cell(row=row_idx, column=1, value=col_a)
        cell_b = ws.cell(row=row_idx, column=2, value=col_b)
        if font_override:
            cell_a.font = font_override
        elif col_a.startswith("  "):
            cell_a.font = CODE_FONT
            cell_b.font = HINT_FONT
        else:
            cell_a.font = BODY_FONT
            cell_b.font = BODY_FONT

    ws.column_dimensions["A"].width = 70
    ws.column_dimensions["B"].width = 65


def generate_template_xlsx() -> bytes:
    """Generate an empty canonical SRS template XLSX with all sheets, validations, and examples.

    Returns bytes for direct download.
    """
    wb = openpyxl.Workbook()
    # Remove default sheet
    wb.remove(wb.active)

    # Instructions
    _build_instructions_sheet(wb)

    # 1. Modelling
    ws = wb.create_sheet("Modelling")
    _style_header_row(ws, MODELLING_HEADERS)
    _add_example_rows(ws, MODELLING_EXAMPLES)
    _auto_width(ws, MODELLING_HEADERS)
    _add_data_validation(ws, 1, VALID_ENTITY_TYPES)  # Entity Type
    _add_data_validation(ws, 5, VALID_FORM_TYPES)     # Form Type

    # 2. Location Hierarchy
    ws = wb.create_sheet("Location Hierarchy")
    _style_header_row(ws, LOCATION_HEADERS)
    _add_example_rows(ws, LOCATION_EXAMPLES)
    _auto_width(ws, LOCATION_HEADERS)

    # 3. Groups
    ws = wb.create_sheet("Groups")
    _style_header_row(ws, GROUPS_HEADERS)
    _add_example_rows(ws, GROUPS_EXAMPLES)
    _auto_width(ws, GROUPS_HEADERS)
    _add_data_validation(ws, 2, ["Yes", "No"])  # Has All Privileges

    # 4. Visit Schedules
    ws = wb.create_sheet("Visit Schedules")
    _style_header_row(ws, VISIT_HEADERS)
    _add_example_rows(ws, VISIT_EXAMPLES)
    _auto_width(ws, VISIT_HEADERS)
    _add_data_validation(ws, 6, ["Reschedule", "Close"])  # On Cancellation

    # 5. Decisions
    ws = wb.create_sheet("Decisions")
    _style_header_row(ws, DECISIONS_HEADERS)
    _add_example_rows(ws, DECISIONS_EXAMPLES)
    _auto_width(ws, DECISIONS_HEADERS)
    _add_data_validation(ws, 5, ["encounter", "enrolment", "registration"])  # Scope

    # 6. Eligibility
    ws = wb.create_sheet("Eligibility")
    _style_header_row(ws, ELIGIBILITY_HEADERS)
    _add_example_rows(ws, ELIGIBILITY_EXAMPLES)
    _auto_width(ws, ELIGIBILITY_HEADERS)

    # 7. Report Cards
    ws = wb.create_sheet("Report Cards")
    _style_header_row(ws, REPORT_CARDS_HEADERS)
    _add_example_rows(ws, REPORT_CARDS_EXAMPLES)
    _auto_width(ws, REPORT_CARDS_HEADERS)
    _add_data_validation(ws, 2, VALID_CARD_TYPES)  # Card Type
    _add_data_validation(ws, 10, ["Yes", "No"])     # Nested

    # 8. "Example Form" — a reference sheet showing all column types with examples.
    #    Users should copy this sheet's header row when creating their own form sheets.
    #    This sheet is NOT parsed as a form (it won't match any Modelling row).
    ws = wb.create_sheet("Example Form")
    _style_header_row(ws, FORM_HEADERS)
    _auto_width(ws, FORM_HEADERS)
    _add_data_validation(ws, 3, VALID_DATA_TYPES)       # Data Type
    _add_data_validation(ws, 4, ["Yes", "No"])           # Mandatory
    _add_data_validation(ws, 6, ["Single", "Multi"])     # Selection Type

    # Add diverse example rows showing all column features with generic field names
    example_note_fill = PatternFill(start_color="F0F0F0", end_color="F0F0F0", fill_type="solid")
    note_row = 2
    note_cell = ws.cell(row=note_row, column=1,
        value="⬇ EXAMPLES ONLY — delete these rows. Copy the header row (row 1) when creating your own form sheets.")
    note_cell.font = Font(name="Calibri", bold=True, italic=True, color="CC6600", size=10)
    for c in range(1, len(FORM_HEADERS) + 1):
        ws.cell(row=note_row, column=c).fill = example_note_fill

    _add_example_rows(ws, [
        # Basic types
        ["Basic Info", "Name", "Text", "Yes", "", "", "", "", "", "", "", ""],
        ["Basic Info", "Date of Birth", "Date", "Yes", "", "", "", "", "", "", "", ""],
        ["Basic Info", "Phone Number", "PhoneNumber", "No", "", "", "", "", "", "", "", ""],
        # Coded — single select with skip logic
        ["Basic Info", "Gender", "Coded", "Yes", "Male; Female; Other", "Single", "", "", "", "", "", ""],
        ["Basic Info", "Other Gender", "Text", "No", "", "", "", "", "", "Gender = Other", "", ""],
        # Coded — multi-select with CONTAINS skip logic
        ["Assessment", "Issues Found", "Coded", "No", "Issue A; Issue B; Issue C; Other; None", "Multi", "", "", "", "", "", ""],
        ["Assessment", "Other Issue Details", "Notes", "No", "", "", "", "", "", "Issues Found CONTAINS Other", "", ""],
        # Numeric with unit and validation range
        ["Measurements", "Score", "Numeric", "Yes", "", "", "points", "0", "100", "", "", ""],
        ["Measurements", "Count", "Numeric", "No", "", "", "", "0", "999", "", "", ""],
        # Show When / Hide When examples
        ["Status", "Status", "Coded", "Yes", "Active; Inactive; Referred", "Single", "", "", "", "", "", ""],
        ["Status", "Referral Reason", "Text", "No", "", "", "", "", "", "Status = Referred", "", ""],
        ["Status", "Inactive Since", "Date", "No", "", "", "", "", "", "Status = Inactive", "", ""],
        # IS EMPTY / IS NOT EMPTY
        ["Follow-up", "Follow-up Needed", "Coded", "Yes", "Yes; No", "Single", "", "", "", "", "", ""],
        ["Follow-up", "Follow-up Date", "Date", "No", "", "", "", "", "", "Follow-up Needed = Yes", "", ""],
        ["Follow-up", "Follow-up Notes", "Notes", "No", "", "", "", "", "", "Follow-up Date IS NOT EMPTY", "", ""],
        # Compound skip logic
        ["Review", "Eligible for Review", "Coded", "No", "Yes; No", "Single", "", "", "", "Score >= 50 AND Status = Active", "", ""],
        # QuestionGroup — repeating group of child fields
        ["Household", "Members", "QuestionGroup", "No", "", "", "", "", "", "", "", ""],
        ["Household", "Member Name", "Text", "Yes", "", "", "", "", "", "", "", "Members"],
        ["Household", "Member Age", "Numeric", "No", "", "", "years", "0", "120", "", "", "Members"],
        ["Household", "Member Relation", "Coded", "No", "Spouse; Child; Parent; Sibling; Other", "Single", "", "", "", "", "", "Members"],
        # Media and special types
        ["Documents", "Photo", "Image", "No", "", "", "", "", "", "", "", ""],
        ["Documents", "GPS Location", "Location", "No", "", "", "", "", "", "", "", ""],
        ["Documents", "Linked Beneficiary", "Subject", "No", "", "", "", "", "", "", "", ""],
        ["Documents", "Duration of Stay", "Duration", "No", "", "", "", "", "", "", "", ""],
    ], start_row=3)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def generate_template_csv() -> bytes:
    """Generate a flat CSV version of the template (Modelling + Form Fields combined).

    Has an extra 'Form Name' column to distinguish which form each field belongs to.
    Returns bytes for direct download.
    """
    buf = io.StringIO()
    writer = csv.writer(buf)

    # Section 1: Modelling
    writer.writerow(["# MODELLING"])
    writer.writerow(MODELLING_HEADERS)
    for row in MODELLING_EXAMPLES:
        writer.writerow(row)
    writer.writerow([])

    # Section 2: Location Hierarchy
    writer.writerow(["# LOCATION HIERARCHY"])
    writer.writerow(LOCATION_HEADERS)
    for row in LOCATION_EXAMPLES:
        writer.writerow(row)
    writer.writerow([])

    # Section 3: Groups
    writer.writerow(["# GROUPS"])
    writer.writerow(GROUPS_HEADERS)
    for row in GROUPS_EXAMPLES:
        writer.writerow(row)
    writer.writerow([])

    # Section 4: Visit Schedules
    writer.writerow(["# VISIT SCHEDULES"])
    writer.writerow(VISIT_HEADERS)
    for row in VISIT_EXAMPLES:
        writer.writerow(row)
    writer.writerow([])

    # Section 5: Decisions
    writer.writerow(["# DECISIONS"])
    writer.writerow(DECISIONS_HEADERS)
    for row in DECISIONS_EXAMPLES:
        writer.writerow(row)
    writer.writerow([])

    # Section 6: Eligibility
    writer.writerow(["# ELIGIBILITY"])
    writer.writerow(ELIGIBILITY_HEADERS)
    for row in ELIGIBILITY_EXAMPLES:
        writer.writerow(row)
    writer.writerow([])

    # Section 7: Report Cards
    writer.writerow(["# REPORT CARDS"])
    writer.writerow(REPORT_CARDS_HEADERS)
    for row in REPORT_CARDS_EXAMPLES:
        writer.writerow(row)
    writer.writerow([])

    # Section 8: Form Fields (flat, with Form Name column)
    writer.writerow(["# FORM FIELDS"])
    writer.writerow(["Form Name"] + FORM_HEADERS)

    return buf.getvalue().encode("utf-8")


def generate_filled_template(srs_data: SRSData) -> bytes:
    """Generate a canonical template pre-filled from existing SRSData.

    Enables round-tripping: upload free-form → download canonical → edit → re-upload.
    """
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    _build_instructions_sheet(wb)

    # 1. Modelling sheet
    ws = wb.create_sheet("Modelling")
    _style_header_row(ws, MODELLING_HEADERS)
    _auto_width(ws, MODELLING_HEADERS)

    row_idx = 2
    # Subject types
    for st in srs_data.subjectTypes:
        name = st.get("name", "Individual")
        st_type = st.get("type", "Person")
        # Find registration form
        reg_form = next(
            (f for f in srs_data.forms if f.formType == "IndividualProfile"
             and (f.subjectTypeName == name or not f.subjectTypeName)),
            None,
        )
        ws.cell(row=row_idx, column=1, value="SubjectType")
        ws.cell(row=row_idx, column=2, value=name)
        ws.cell(row=row_idx, column=3, value=st_type)
        ws.cell(row=row_idx, column=4, value=reg_form.name if reg_form else f"{name} Registration")
        ws.cell(row=row_idx, column=5, value="IndividualProfile")
        row_idx += 1

    # Programs
    for prog in srs_data.programs:
        prog_name = prog.get("name", prog) if isinstance(prog, dict) else prog
        colour = prog.get("colour", "") if isinstance(prog, dict) else ""
        enrol_form = next(
            (f for f in srs_data.forms if f.formType == "ProgramEnrolment" and f.programName == prog_name),
            None,
        )
        ws.cell(row=row_idx, column=1, value="Program")
        ws.cell(row=row_idx, column=2, value=prog_name)
        ws.cell(row=row_idx, column=3, value="")
        ws.cell(row=row_idx, column=4, value=enrol_form.name if enrol_form else f"{prog_name} Enrolment")
        ws.cell(row=row_idx, column=5, value="ProgramEnrolment")
        ws.cell(row=row_idx, column=6, value=colour)
        row_idx += 1

    # Encounter types (from forms)
    for form_def in srs_data.forms:
        if form_def.formType in ("ProgramEncounter", "Encounter"):
            et_name = form_def.encounterTypeName or form_def.name
            parent = form_def.programName or ""
            ws.cell(row=row_idx, column=1, value="EncounterType")
            ws.cell(row=row_idx, column=2, value=et_name)
            ws.cell(row=row_idx, column=3, value=parent)
            ws.cell(row=row_idx, column=4, value=form_def.name)
            ws.cell(row=row_idx, column=5, value=form_def.formType)
            row_idx += 1

    # 2. Location Hierarchy
    ws = wb.create_sheet("Location Hierarchy")
    _style_header_row(ws, LOCATION_HEADERS)
    _auto_width(ws, LOCATION_HEADERS)
    if srs_data.addressLevelTypes:
        for i, alt in enumerate(srs_data.addressLevelTypes, 2):
            ws.cell(row=i, column=1, value=alt.get("name", ""))
            ws.cell(row=i, column=2, value=alt.get("level", ""))
            ws.cell(row=i, column=3, value=alt.get("parent", ""))

    # 3. Groups
    ws = wb.create_sheet("Groups")
    _style_header_row(ws, GROUPS_HEADERS)
    _auto_width(ws, GROUPS_HEADERS)
    for i, group_name in enumerate(srs_data.groups, 2):
        ws.cell(row=i, column=1, value=group_name)
        ws.cell(row=i, column=2, value="Yes" if group_name == "Everyone" else "No")

    # 4. Visit Schedules
    ws = wb.create_sheet("Visit Schedules")
    _style_header_row(ws, VISIT_HEADERS)
    _auto_width(ws, VISIT_HEADERS)
    if srs_data.visitSchedules:
        for i, vs in enumerate(srs_data.visitSchedules, 2):
            ws.cell(row=i, column=1, value=vs.get("trigger", vs.get("triggerForm", "")))
            ws.cell(row=i, column=2, value=vs.get("schedule_encounter", vs.get("encounterType", "")))
            ws.cell(row=i, column=3, value=vs.get("visit_name", ""))
            ws.cell(row=i, column=4, value=vs.get("due_days", vs.get("dueDays", "")))
            ws.cell(row=i, column=5, value=vs.get("overdue_days", vs.get("overdueDays", "")))
            ws.cell(row=i, column=6, value=vs.get("on_cancellation", ""))

    # 5. Form sheets — one per form (skip auto-generated cancellation/exit)
    for form_def in srs_data.forms:
        if form_def.formType in ("ProgramEncounterCancellation", "IndividualEncounterCancellation", "ProgramExit"):
            continue  # These are auto-generated
        ws = wb.create_sheet(form_def.name[:31])  # Excel limits sheet names to 31 chars
        _style_header_row(ws, FORM_HEADERS)
        _auto_width(ws, FORM_HEADERS)
        _add_data_validation(ws, 3, VALID_DATA_TYPES)
        _add_data_validation(ws, 4, ["Yes", "No"])
        _add_data_validation(ws, 6, ["Single", "Multi"])

        row_idx = 2
        for group in form_def.groups:
            for field in group.fields:
                ws.cell(row=row_idx, column=1, value=group.name)
                ws.cell(row=row_idx, column=2, value=field.name)
                ws.cell(row=row_idx, column=3, value=field.dataType)
                ws.cell(row=row_idx, column=4, value="Yes" if field.mandatory else "No")
                if field.options:
                    ws.cell(row=row_idx, column=5, value="; ".join(field.options))
                if field.type:
                    sel = "Multi" if field.type.lower() in ("multiselect", "multi") else "Single"
                    ws.cell(row=row_idx, column=6, value=sel)
                if field.unit:
                    ws.cell(row=row_idx, column=7, value=field.unit)
                if field.lowAbsolute is not None:
                    ws.cell(row=row_idx, column=8, value=field.lowAbsolute)
                if field.highAbsolute is not None:
                    ws.cell(row=row_idx, column=9, value=field.highAbsolute)

                # Extract showWhen/hideWhen from keyValues
                if field.keyValues:
                    for kv in field.keyValues:
                        if kv.get("key") == "showWhen":
                            ws.cell(row=row_idx, column=10, value=kv.get("value", ""))
                        elif kv.get("key") == "hideWhen":
                            ws.cell(row=row_idx, column=11, value=kv.get("value", ""))
                        elif kv.get("key") == "qgParent":
                            ws.cell(row=row_idx, column=12, value=kv.get("value", ""))

                row_idx += 1

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
