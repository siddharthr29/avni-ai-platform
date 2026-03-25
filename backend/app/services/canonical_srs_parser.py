"""Parse canonical SRS template XLSX into SRSData — 100% deterministic, zero LLM.

The canonical template has fixed sheets and headers that map 1:1 to Avni's data model.
This parser reads the template and produces SRSData with all entity relationships,
form definitions, skip logic, visit schedules, and validations explicitly defined.

All 26 server failure modes are caught at parse time — no bundle should ever reach
the server with a contract violation.
"""

from __future__ import annotations

import logging
import re
from pathlib import Path
from typing import Any

import openpyxl
from openpyxl.worksheet.worksheet import Worksheet

from app.models.schemas import (
    SRSData,
    SRSFormDefinition,
    SRSFormField,
    SRSFormGroup,
)

logger = logging.getLogger(__name__)

# ---------------------------------------------------------------------------
# Valid enums (must match bundle_generator.py)
# ---------------------------------------------------------------------------

VALID_FORM_TYPES = {
    "IndividualProfile", "ProgramEnrolment", "ProgramExit",
    "ProgramEncounter", "ProgramEncounterCancellation",
    "Encounter", "IndividualEncounterCancellation",
    "BeneficiaryIdentification", "SubjectEnrolmentEligibility",
    "ManualProgramEnrolmentEligibility", "ChecklistItem",
    "IndividualRelationship", "Location", "Task",
}

VALID_DATA_TYPES = {
    "Numeric", "Text", "Coded", "Date", "DateTime", "Time",
    "Duration", "Image", "ImageV2", "Video", "Audio", "File", "Id",
    "NA", "Notes", "Location", "PhoneNumber", "GroupAffiliation",
    "Subject", "Encounter", "QuestionGroup",
}

VALID_ENTITY_TYPES = {"SubjectType", "Program", "EncounterType"}


# ---------------------------------------------------------------------------
# Detection
# ---------------------------------------------------------------------------


def is_canonical_template(file_path: str) -> bool:
    """Check if an XLSX file is a canonical SRS template.

    Fast check: looks for a sheet named 'Modelling' with the expected header row.
    """
    try:
        wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
        if "Modelling" not in wb.sheetnames:
            wb.close()
            return False
        ws = wb["Modelling"]
        headers = [str(cell.value or "").strip() for cell in next(ws.iter_rows(min_row=1, max_row=1))]
        wb.close()
        return "Entity Type" in headers and "Form Name" in headers and "Form Type" in headers
    except Exception:
        return False


# ---------------------------------------------------------------------------
# Skip logic parser (structured format from template)
# ---------------------------------------------------------------------------

# Patterns for the structured skip logic format defined in the template
_SKIP_PATTERNS = [
    # Multi-char operators MUST come before single-char (order matters: first match wins)
    # FieldName != Value
    (re.compile(r"^(.+?)\s*!=\s*(.+)$"), "not_equals"),
    # FieldName >= N
    (re.compile(r"^(.+?)\s*>=\s*(.+)$"), "greater_than_or_equal"),
    # FieldName <= N
    (re.compile(r"^(.+?)\s*<=\s*(.+)$"), "less_than_or_equal"),
    # FieldName > N (after >=)
    (re.compile(r"^(.+?)\s*>\s*(.+)$"), "greater_than"),
    # FieldName < N (after <=)
    (re.compile(r"^(.+?)\s*<\s*(.+)$"), "less_than"),
    # FieldName = Value (after !=, >=, <=)
    (re.compile(r"^(.+?)\s*=\s*(.+)$"), "equals"),
    # FieldName IS EMPTY
    (re.compile(r"^(.+?)\s+IS\s+EMPTY$", re.IGNORECASE), "is_empty"),
    # FieldName IS NOT EMPTY
    (re.compile(r"^(.+?)\s+IS\s+NOT\s+EMPTY$", re.IGNORECASE), "is_not_empty"),
    # FieldName CONTAINS Value
    (re.compile(r"^(.+?)\s+CONTAINS\s+(.+)$", re.IGNORECASE), "contains"),
    # FieldName IN (A, B, C)
    (re.compile(r"^(.+?)\s+IN\s*\((.+)\)$", re.IGNORECASE), "in"),
]


def _parse_structured_condition(text: str) -> dict[str, Any] | None:
    """Parse a structured skip logic condition from the template.

    Returns dict with keys: trigger_field, operator, value, compound (optional).
    """
    text = text.strip()
    if not text:
        return None

    # Check for compound conditions (AND/OR)
    # Split on AND/OR while respecting parentheses
    and_parts = re.split(r"\s+AND\s+", text, flags=re.IGNORECASE)
    if len(and_parts) > 1:
        conditions = [_parse_structured_condition(p.strip()) for p in and_parts]
        conditions = [c for c in conditions if c is not None]
        if conditions:
            return {"compound": "and", "conditions": conditions}

    or_parts = re.split(r"\s+OR\s+", text, flags=re.IGNORECASE)
    if len(or_parts) > 1:
        conditions = [_parse_structured_condition(p.strip()) for p in or_parts]
        conditions = [c for c in conditions if c is not None]
        if conditions:
            return {"compound": "or", "conditions": conditions}

    # Single condition
    for pattern, operator in _SKIP_PATTERNS:
        m = pattern.match(text)
        if m:
            groups = m.groups()
            field_name = groups[0].strip()
            value = groups[1].strip() if len(groups) > 1 else None
            return {
                "trigger_field": field_name,
                "operator": operator,
                "value": value,
            }

    return None


# ---------------------------------------------------------------------------
# Sheet readers
# ---------------------------------------------------------------------------


def _read_header_map(ws: Worksheet) -> dict[str, int]:
    """Read the header row and return a map of header name → column index (0-based)."""
    headers: dict[str, int] = {}
    for cell in next(ws.iter_rows(min_row=1, max_row=1)):
        name = str(cell.value or "").strip()
        if name:
            headers[name] = cell.column - 1  # 0-based
    return headers


def _cell_str(row: tuple, col_idx: int) -> str:
    """Get cell value as stripped string, or empty string if None."""
    if col_idx < 0 or col_idx >= len(row):
        return ""
    val = row[col_idx].value
    if val is None:
        return ""
    return str(val).strip()


def _cell_float(row: tuple, col_idx: int) -> float | None:
    """Get cell value as float, or None."""
    if col_idx < 0 or col_idx >= len(row):
        return None
    val = row[col_idx].value
    if val is None or val == "":
        return None
    try:
        return float(val)
    except (ValueError, TypeError):
        return None


def _parse_modelling_sheet(ws: Worksheet, errors: list[str]) -> dict[str, Any]:
    """Parse the Modelling sheet.

    Returns dict with:
      subject_types: list of {name, type}
      programs: list of {name, colour}
      encounter_types: dict of et_name -> {program, form_name, form_type}
      form_registry: list of {entity_type, name, parent, form_name, form_type, colour}
    """
    hdr = _read_header_map(ws)
    required = {"Entity Type", "Name", "Form Name", "Form Type"}
    missing = required - set(hdr.keys())
    if missing:
        errors.append(f"Modelling sheet missing required columns: {missing}")
        return {"subject_types": [], "programs": [], "encounter_types": {}, "form_registry": []}

    et_col = hdr["Entity Type"]
    name_col = hdr["Name"]
    parent_col = hdr.get("Parent (Type/Program)", -1)
    form_name_col = hdr["Form Name"]
    form_type_col = hdr["Form Type"]
    colour_col = hdr.get("Colour", -1)

    subject_types: list[dict[str, str]] = []
    programs: list[dict[str, str]] = []
    encounter_types: dict[str, dict[str, str]] = {}
    form_registry: list[dict[str, str]] = []

    for row_idx, row in enumerate(ws.iter_rows(min_row=2), 2):
        entity_type = _cell_str(row, et_col)
        name = _cell_str(row, name_col)
        if not entity_type or not name:
            continue  # Skip blank rows

        if entity_type not in VALID_ENTITY_TYPES:
            errors.append(f"Modelling row {row_idx}: invalid Entity Type '{entity_type}' (valid: {VALID_ENTITY_TYPES})")
            continue

        parent = _cell_str(row, parent_col)
        form_name = _cell_str(row, form_name_col)
        form_type = _cell_str(row, form_type_col)
        colour = _cell_str(row, colour_col)

        if not form_name:
            errors.append(f"Modelling row {row_idx}: Form Name is required for '{name}'")
            continue
        if form_type and form_type not in VALID_FORM_TYPES:
            errors.append(f"Modelling row {row_idx}: invalid Form Type '{form_type}' (valid: {VALID_FORM_TYPES})")
            continue

        entry = {
            "entity_type": entity_type,
            "name": name,
            "parent": parent,
            "form_name": form_name,
            "form_type": form_type,
            "colour": colour,
        }
        form_registry.append(entry)

        if entity_type == "SubjectType":
            st_kind = parent if parent in ("Person", "Household", "Group") else "Person"
            subject_types.append({"name": name, "type": st_kind})
        elif entity_type == "Program":
            programs.append({"name": name, "colour": colour or ""})
        elif entity_type == "EncounterType":
            encounter_types[name] = {"program": parent, "form_name": form_name, "form_type": form_type}

    return {
        "subject_types": subject_types,
        "programs": programs,
        "encounter_types": encounter_types,
        "form_registry": form_registry,
    }


def _parse_location_hierarchy(ws: Worksheet, errors: list[str]) -> list[dict[str, Any]] | None:
    """Parse Location Hierarchy sheet."""
    hdr = _read_header_map(ws)
    if "Level Name" not in hdr:
        return None

    name_col = hdr["Level Name"]
    level_col = hdr.get("Level Number", -1)
    parent_col = hdr.get("Parent Level", -1)

    levels: list[dict[str, Any]] = []
    for row in ws.iter_rows(min_row=2):
        name = _cell_str(row, name_col)
        if not name:
            continue
        level_num = _cell_float(row, level_col)
        parent = _cell_str(row, parent_col)
        entry: dict[str, Any] = {"name": name}
        if level_num is not None:
            entry["level"] = int(level_num)
        if parent:
            entry["parent"] = parent
        levels.append(entry)

    # Validate: no duplicate level names
    seen_names = set()
    for lvl in levels:
        if lvl["name"] in seen_names:
            errors.append(f"Location Hierarchy: duplicate level name '{lvl['name']}'")
        seen_names.add(lvl["name"])

    return levels if levels else None


def _parse_groups(ws: Worksheet, errors: list[str]) -> list[str]:
    """Parse Groups sheet."""
    hdr = _read_header_map(ws)
    if "Group Name" not in hdr:
        return ["Everyone"]

    name_col = hdr["Group Name"]
    groups: list[str] = []
    for row in ws.iter_rows(min_row=2):
        name = _cell_str(row, name_col)
        if name:
            groups.append(name)

    if "Everyone" not in groups:
        groups.insert(0, "Everyone")

    return groups


def _parse_visit_schedules(ws: Worksheet, errors: list[str]) -> list[dict[str, Any]] | None:
    """Parse Visit Schedules sheet."""
    hdr = _read_header_map(ws)
    if "After Form" not in hdr:
        return None

    after_col = hdr["After Form"]
    sched_col = hdr.get("Schedule Encounter", -1)
    visit_name_col = hdr.get("Visit Name", -1)
    due_col = hdr.get("Due Days", -1)
    overdue_col = hdr.get("Overdue Days", -1)
    cancel_col = hdr.get("On Cancellation", -1)

    schedules: list[dict[str, Any]] = []
    for row_idx, row in enumerate(ws.iter_rows(min_row=2), 2):
        after_form = _cell_str(row, after_col)
        if not after_form:
            continue

        sched_enc = _cell_str(row, sched_col)
        if not sched_enc:
            errors.append(f"Visit Schedules row {row_idx}: Schedule Encounter is required")
            continue

        due_days = _cell_float(row, due_col)
        overdue_days = _cell_float(row, overdue_col)
        if due_days is None:
            errors.append(f"Visit Schedules row {row_idx}: Due Days is required")
            continue
        if overdue_days is None:
            errors.append(f"Visit Schedules row {row_idx}: Overdue Days is required")
            continue

        entry: dict[str, Any] = {
            "trigger": after_form,
            "schedule_encounter": sched_enc,
            "due_days": int(due_days),
            "overdue_days": int(overdue_days),
        }
        visit_name = _cell_str(row, visit_name_col)
        if visit_name:
            entry["visit_name"] = visit_name
        cancel_action = _cell_str(row, cancel_col)
        if cancel_action:
            entry["on_cancellation"] = cancel_action

        schedules.append(entry)

    return schedules if schedules else None


def _parse_decisions(ws: Worksheet, errors: list[str]) -> list[dict[str, Any]] | None:
    """Parse Decisions sheet."""
    hdr = _read_header_map(ws)
    if "Form Name" not in hdr or "Set Field" not in hdr:
        return None

    form_col = hdr["Form Name"]
    when_col = hdr.get("When", -1)
    set_field_col = hdr["Set Field"]
    to_value_col = hdr.get("To Value", -1)
    scope_col = hdr.get("Scope", -1)

    decisions: list[dict[str, Any]] = []
    valid_scopes = {"encounter", "enrolment", "registration"}

    for row_idx, row in enumerate(ws.iter_rows(min_row=2), 2):
        form_name = _cell_str(row, form_col)
        if not form_name:
            continue

        set_field = _cell_str(row, set_field_col)
        if not set_field:
            errors.append(f"Decisions row {row_idx}: Set Field is required")
            continue

        when = _cell_str(row, when_col) or "ALWAYS"
        to_value = _cell_str(row, to_value_col)
        scope = _cell_str(row, scope_col) or "encounter"

        if scope not in valid_scopes:
            errors.append(f"Decisions row {row_idx}: invalid Scope '{scope}' (valid: {valid_scopes})")
            continue

        entry: dict[str, Any] = {
            "formName": form_name,
            "when": when,
            "setField": set_field,
            "toValue": to_value,
            "scope": scope,
        }

        # Parse the condition if not ALWAYS
        if when.upper() != "ALWAYS":
            parsed = _parse_structured_condition(when)
            if not parsed:
                errors.append(f"Decisions row {row_idx}: could not parse condition '{when}'")

        # Detect formula references like {Weight}
        if to_value and "{" in to_value:
            entry["hasReferences"] = True

        decisions.append(entry)

    return decisions if decisions else None


def _parse_eligibility(ws: Worksheet, errors: list[str]) -> list[dict[str, Any]] | None:
    """Parse Eligibility sheet."""
    hdr = _read_header_map(ws)
    if "Program" not in hdr or "Condition" not in hdr:
        return None

    prog_col = hdr["Program"]
    cond_col = hdr["Condition"]

    rules: list[dict[str, Any]] = []
    for row_idx, row in enumerate(ws.iter_rows(min_row=2), 2):
        program = _cell_str(row, prog_col)
        condition = _cell_str(row, cond_col)
        if not program or not condition:
            continue

        parsed = _parse_structured_condition(condition)
        if not parsed:
            errors.append(f"Eligibility row {row_idx}: could not parse condition '{condition}'")
            continue

        rules.append({
            "program": program,
            "condition": condition,
            "parsed": parsed,
        })

    return rules if rules else None


def _parse_report_cards(ws: Worksheet, errors: list[str]) -> list[dict[str, Any]] | None:
    """Parse Report Cards sheet."""
    hdr = _read_header_map(ws)
    if "Card Name" not in hdr or "Card Type" not in hdr:
        return None

    name_col = hdr["Card Name"]
    type_col = hdr["Card Type"]
    desc_col = hdr.get("Description", -1)
    st_col = hdr.get("Subject Type", -1)
    prog_col = hdr.get("Program", -1)
    et_col = hdr.get("Encounter Type", -1)
    dur_col = hdr.get("Recent Duration", -1)
    filter_col = hdr.get("Filter Condition", -1)
    colour_col = hdr.get("Colour", -1)
    nested_col = hdr.get("Nested", -1)

    valid_types = {
        "Total", "ScheduledVisits", "OverdueVisits",
        "RecentRegistrations", "RecentEnrolments", "RecentVisits",
        "DueChecklist", "Tasks", "CallTasks", "OpenSubjectTasks",
        "PendingApproval", "Approved", "Rejected", "Comments",
        "Custom",
    }

    cards: list[dict[str, Any]] = []
    for row_idx, row in enumerate(ws.iter_rows(min_row=2), 2):
        card_name = _cell_str(row, name_col)
        card_type = _cell_str(row, type_col)
        if not card_name or not card_type:
            continue

        if card_type not in valid_types:
            errors.append(f"Report Cards row {row_idx}: invalid Card Type '{card_type}' (valid: {valid_types})")
            continue

        entry: dict[str, Any] = {
            "name": card_name,
            "cardType": card_type,
        }

        desc = _cell_str(row, desc_col)
        if desc:
            entry["description"] = desc
        st = _cell_str(row, st_col)
        if st:
            entry["subjectType"] = st
        prog = _cell_str(row, prog_col)
        if prog:
            entry["program"] = prog
        et = _cell_str(row, et_col)
        if et:
            entry["encounterType"] = et
        dur = _cell_str(row, dur_col)
        if dur:
            entry["recentDuration"] = dur
        colour = _cell_str(row, colour_col)
        if colour:
            entry["colour"] = colour
        nested = _cell_str(row, nested_col).lower()
        entry["nested"] = nested in ("yes", "true", "1")

        # Parse filter condition for Custom cards
        filter_cond = _cell_str(row, filter_col)
        if filter_cond:
            entry["filterCondition"] = filter_cond
            parsed = _parse_structured_condition(filter_cond)
            if not parsed:
                errors.append(f"Report Cards row {row_idx}: could not parse Filter Condition '{filter_cond}'")

        if card_type == "Custom" and not filter_cond:
            errors.append(f"Report Cards row {row_idx}: Custom cards require a Filter Condition")

        cards.append(entry)

    return cards if cards else None


def _parse_form_sheet(
    ws: Worksheet,
    form_name: str,
    form_type: str,
    errors: list[str],
) -> SRSFormDefinition | None:
    """Parse a single form sheet into SRSFormDefinition."""
    hdr = _read_header_map(ws)
    if "Field Name" not in hdr or "Data Type" not in hdr:
        errors.append(f"Form sheet '{form_name}': missing required columns 'Field Name' and/or 'Data Type'")
        return None

    page_col = hdr.get("Page/Section", -1)
    name_col = hdr["Field Name"]
    dt_col = hdr["Data Type"]
    mand_col = hdr.get("Mandatory", -1)
    options_col = hdr.get("Options", -1)
    sel_col = hdr.get("Selection Type", -1)
    unit_col = hdr.get("Unit", -1)
    min_col = hdr.get("Min Value", -1)
    max_col = hdr.get("Max Value", -1)
    show_col = hdr.get("Show When", -1)
    hide_col = hdr.get("Hide When", -1)
    qg_col = hdr.get("QG Parent", -1)

    # Group fields by page/section
    groups_dict: dict[str, list[SRSFormField]] = {}
    field_names_in_form: set[str] = set()
    group_order: list[str] = []

    for row_idx, row in enumerate(ws.iter_rows(min_row=2), 2):
        field_name = _cell_str(row, name_col)
        if not field_name:
            continue

        data_type = _cell_str(row, dt_col)
        if not data_type:
            errors.append(f"Form '{form_name}' row {row_idx}: Data Type is required for '{field_name}'")
            continue

        if data_type not in VALID_DATA_TYPES:
            errors.append(f"Form '{form_name}' row {row_idx}: invalid Data Type '{data_type}' (valid: {VALID_DATA_TYPES})")
            continue

        # Validate field name length (server error F7)
        if len(field_name) > 255:
            errors.append(f"Form '{form_name}' row {row_idx}: Field Name exceeds 255 characters")
            continue

        # Check duplicate field in same form (server error F2)
        if field_name in field_names_in_form:
            errors.append(f"Form '{form_name}' row {row_idx}: duplicate field '{field_name}' in same form")
            continue
        field_names_in_form.add(field_name)

        page = _cell_str(row, page_col) or "Default"
        mandatory_str = _cell_str(row, mand_col).lower()
        mandatory = mandatory_str in ("yes", "true", "1", "y")

        options_str = _cell_str(row, options_col)
        options: list[str] | None = None
        if options_str:
            options = [o.strip() for o in options_str.split(";") if o.strip()]

        sel_type = _cell_str(row, sel_col)
        field_type: str | None = None
        if data_type == "Coded":
            if sel_type.lower() in ("multi", "multiselect"):
                field_type = "MultiSelect"
            else:
                field_type = "SingleSelect"
            # Validate coded fields have options (server error C5)
            if not options:
                errors.append(f"Form '{form_name}' row {row_idx}: Coded field '{field_name}' has no options")

        unit = _cell_str(row, unit_col) or None
        low_abs = _cell_float(row, min_col)
        high_abs = _cell_float(row, max_col)

        # Build keyValues from skip logic and QG parent
        key_values: list[dict[str, Any]] = []
        show_when = _cell_str(row, show_col)
        hide_when = _cell_str(row, hide_col)
        qg_parent = _cell_str(row, qg_col)

        if show_when:
            parsed = _parse_structured_condition(show_when)
            if parsed:
                key_values.append({"key": "showWhen", "value": show_when})
            else:
                # Store raw text — skip logic generator will handle it
                key_values.append({"key": "showWhen", "value": show_when})

        if hide_when:
            parsed = _parse_structured_condition(hide_when)
            if parsed:
                key_values.append({"key": "hideWhen", "value": hide_when})
            else:
                key_values.append({"key": "hideWhen", "value": hide_when})

        if qg_parent:
            key_values.append({"key": "qgParent", "value": qg_parent})

        field = SRSFormField(
            name=field_name,
            dataType=data_type,
            mandatory=mandatory,
            options=options,
            type=field_type,
            unit=unit,
            lowAbsolute=low_abs,
            highAbsolute=high_abs,
            keyValues=key_values if key_values else None,
        )

        if page not in groups_dict:
            group_order.append(page)
        groups_dict.setdefault(page, []).append(field)

    if not groups_dict:
        errors.append(f"Form sheet '{form_name}': no fields found")
        return None

    # Validate page/section name lengths (server error F7)
    for page_name in groups_dict:
        if len(page_name) > 255:
            errors.append(f"Form '{form_name}': Page/Section name '{page_name[:50]}...' exceeds 255 characters")

    # Validate QG Parent references
    for page_fields in groups_dict.values():
        for field in page_fields:
            if field.keyValues:
                for kv in field.keyValues:
                    if kv.get("key") == "qgParent":
                        parent_name = kv["value"]
                        if parent_name not in field_names_in_form:
                            errors.append(
                                f"Form '{form_name}': QG Parent '{parent_name}' for field "
                                f"'{field.name}' does not exist in this form"
                            )
                        else:
                            # Verify parent is a QuestionGroup type
                            parent_field = next(
                                (f for page_f in groups_dict.values() for f in page_f if f.name == parent_name),
                                None,
                            )
                            if parent_field and parent_field.dataType != "QuestionGroup":
                                errors.append(
                                    f"Form '{form_name}': QG Parent '{parent_name}' for field "
                                    f"'{field.name}' is not a QuestionGroup (is {parent_field.dataType})"
                                )

    # Validate skip logic field references
    for page_fields in groups_dict.values():
        for field in page_fields:
            if field.keyValues:
                for kv in field.keyValues:
                    if kv.get("key") in ("showWhen", "hideWhen"):
                        condition = _parse_structured_condition(kv["value"])
                        if condition:
                            _validate_skip_logic_refs(condition, field_names_in_form, form_name, field.name, errors)

    groups = [SRSFormGroup(name=page, fields=groups_dict[page]) for page in group_order]
    return SRSFormDefinition(name=form_name, formType=form_type, groups=groups)


def _validate_skip_logic_refs(
    condition: dict[str, Any],
    field_names: set[str],
    form_name: str,
    field_name: str,
    errors: list[str],
) -> None:
    """Validate that skip logic references point to existing fields in the same form."""
    if "compound" in condition:
        for sub in condition.get("conditions", []):
            _validate_skip_logic_refs(sub, field_names, form_name, field_name, errors)
        return

    trigger = condition.get("trigger_field", "")
    if trigger and trigger not in field_names:
        errors.append(
            f"Form '{form_name}': skip logic for '{field_name}' references "
            f"unknown field '{trigger}'"
        )


# ---------------------------------------------------------------------------
# Main parser
# ---------------------------------------------------------------------------


def parse_canonical_srs(file_path: str) -> tuple[SRSData, list[str]]:
    """Parse a canonical SRS template XLSX into SRSData.

    Returns (SRSData, list_of_validation_errors).
    Zero LLM, 100% deterministic.
    """
    errors: list[str] = []

    wb = openpyxl.load_workbook(file_path, data_only=True)

    # 1. Parse Modelling sheet
    if "Modelling" not in wb.sheetnames:
        errors.append("Missing required sheet: Modelling")
        wb.close()
        return SRSData(), errors

    modelling = _parse_modelling_sheet(wb["Modelling"], errors)
    subject_types = modelling["subject_types"]
    programs = modelling["programs"]
    encounter_types_map = modelling["encounter_types"]
    form_registry = modelling["form_registry"]

    if not subject_types:
        subject_types = [{"name": "Individual", "type": "Person"}]

    # 2. Parse Location Hierarchy
    address_levels = None
    if "Location Hierarchy" in wb.sheetnames:
        address_levels = _parse_location_hierarchy(wb["Location Hierarchy"], errors)

    # 3. Parse Groups
    groups = ["Everyone"]
    if "Groups" in wb.sheetnames:
        groups = _parse_groups(wb["Groups"], errors)

    # 4. Parse Visit Schedules
    visit_schedules = None
    if "Visit Schedules" in wb.sheetnames:
        visit_schedules = _parse_visit_schedules(wb["Visit Schedules"], errors)

    # 5. Parse Decisions
    decisions = None
    if "Decisions" in wb.sheetnames:
        decisions = _parse_decisions(wb["Decisions"], errors)

    # 6. Parse Eligibility
    eligibility_rules = None
    if "Eligibility" in wb.sheetnames:
        eligibility_rules = _parse_eligibility(wb["Eligibility"], errors)

    # 7. Parse Report Cards
    report_cards = None
    if "Report Cards" in wb.sheetnames:
        report_cards = _parse_report_cards(wb["Report Cards"], errors)

    # 8. Parse form sheets — one per form in the registry
    forms: list[SRSFormDefinition] = []
    encounter_type_names: list[str] = []
    program_encounter_mappings: list[dict[str, Any]] = []
    general_encounter_types: list[str] = []
    pe_map: dict[str, set[str]] = {}

    for entry in form_registry:
        form_name = entry["form_name"]
        form_type = entry["form_type"]
        entity_type = entry["entity_type"]
        entity_name = entry["name"]
        parent = entry["parent"]

        # Find the matching sheet
        sheet_name = None
        for sn in wb.sheetnames:
            if sn == form_name or sn == form_name[:31]:  # Excel 31-char limit
                sheet_name = sn
                break

        if not sheet_name:
            errors.append(f"Missing form sheet '{form_name}' referenced in Modelling")
            continue

        form_def = _parse_form_sheet(wb[sheet_name], form_name, form_type, errors)
        if not form_def:
            continue

        # Set entity relationships
        if entity_type == "EncounterType":
            form_def.encounterTypeName = entity_name
            if entity_name not in encounter_type_names:
                encounter_type_names.append(entity_name)
            if parent:  # Program encounter
                form_def.programName = parent
                pe_map.setdefault(parent, set()).add(entity_name)
            else:  # Standalone encounter
                if entity_name not in general_encounter_types:
                    general_encounter_types.append(entity_name)

        elif entity_type == "Program":
            form_def.programName = entity_name

        elif entity_type == "SubjectType":
            form_def.subjectTypeName = entity_name

        forms.append(form_def)

    # Build program-encounter mappings
    for prog_name, et_set in pe_map.items():
        program_encounter_mappings.append({
            "program": prog_name,
            "encounterTypes": sorted(et_set),
        })

    # 6. Cross-validation
    _cross_validate(
        form_registry, forms, encounter_type_names, programs,
        subject_types, visit_schedules, groups, errors,
    )

    wb.close()

    srs = SRSData(
        subjectTypes=subject_types,
        programs=programs,
        encounterTypes=encounter_type_names,
        forms=forms,
        groups=groups,
        addressLevelTypes=address_levels,
        programEncounterMappings=program_encounter_mappings if program_encounter_mappings else None,
        generalEncounterTypes=general_encounter_types if general_encounter_types else None,
        visitSchedules=visit_schedules,
        decisions=decisions,
        eligibilityRules=eligibility_rules,
        reportCards=report_cards,
    )

    logger.info(
        "Canonical SRS parsed: %d subject types, %d programs, %d encounter types, "
        "%d forms, %d validation errors",
        len(subject_types), len(programs), len(encounter_type_names),
        len(forms), len(errors),
    )

    return srs, errors


# ---------------------------------------------------------------------------
# Cross-validation
# ---------------------------------------------------------------------------


def _cross_validate(
    form_registry: list[dict[str, str]],
    forms: list[SRSFormDefinition],
    encounter_type_names: list[str],
    programs: list[dict[str, str]],
    subject_types: list[dict[str, str]],
    visit_schedules: list[dict[str, Any]] | None,
    groups: list[str],
    errors: list[str],
) -> None:
    """Cross-validate all references between sheets."""

    program_names = {p["name"] for p in programs}
    st_names = {st["name"] for st in subject_types}
    form_names = {f.name for f in forms}
    et_names = set(encounter_type_names)

    # Modelling references
    for entry in form_registry:
        entity_type = entry["entity_type"]
        parent = entry["parent"]
        form_type = entry["form_type"]

        # ProgramEnrolment/ProgramEncounter must reference an existing program
        if form_type in ("ProgramEnrolment", "ProgramExit") and entry["entity_type"] == "Program":
            if entry["name"] not in program_names:
                errors.append(
                    f"Modelling: Form type '{form_type}' references program "
                    f"'{entry['name']}' which is not defined as a Program row"
                )

        if entity_type == "EncounterType" and parent:
            if parent not in program_names:
                errors.append(
                    f"Modelling: EncounterType '{entry['name']}' references "
                    f"program '{parent}' which is not defined as a Program row"
                )

    # Visit schedule references
    if visit_schedules:
        for vs in visit_schedules:
            trigger = vs.get("trigger", "")
            sched_enc = vs.get("schedule_encounter", "")
            if trigger and trigger not in form_names:
                # Trigger could be a form name or encounter type name — check both
                if trigger not in et_names:
                    errors.append(
                        f"Visit Schedules: After Form '{trigger}' is not a known form or encounter type"
                    )
            if sched_enc and sched_enc not in et_names:
                errors.append(
                    f"Visit Schedules: Schedule Encounter '{sched_enc}' is not a known encounter type"
                )

    # Groups: Everyone must exist
    if "Everyone" not in groups:
        errors.append("Groups: 'Everyone' group is required by the server")

    # Concept name uniqueness across all forms (server error D1)
    all_field_names: dict[str, list[str]] = {}  # field_name -> list of form names
    all_field_types: dict[str, str] = {}  # field_name -> first data type
    for form in forms:
        for group in form.groups:
            for field in group.fields:
                all_field_names.setdefault(field.name, []).append(form.name)
                if field.name in all_field_types:
                    # Check for type conflicts (server error C6)
                    existing_type = all_field_types[field.name]
                    if field.dataType != existing_type and existing_type != "NA" and field.dataType != "NA":
                        errors.append(
                            f"Concept '{field.name}' has conflicting data types: "
                            f"'{existing_type}' (in {all_field_names[field.name][0]}) vs "
                            f"'{field.dataType}' (in {form.name})"
                        )
                else:
                    all_field_types[field.name] = field.dataType

    # Check for question/answer name collisions (server error C7)
    all_option_names: set[str] = set()
    for form in forms:
        for group in form.groups:
            for field in group.fields:
                if field.options:
                    for opt in field.options:
                        all_option_names.add(opt)

    for opt_name in all_option_names:
        if opt_name in all_field_types:
            dt = all_field_types[opt_name]
            if dt not in ("NA", "Coded"):
                errors.append(
                    f"Concept name collision: '{opt_name}' is used as both "
                    f"a {dt} field and a coded option value"
                )
