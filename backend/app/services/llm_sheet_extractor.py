"""LLM-powered sheet extractor for Avni scoping sheets.

Reads any XLSX/CSV file(s), sends sheet metadata to a fast LLM (Groq/Cerebras),
and gets back structured modelling data (subject types, programs, encounters,
form classifications) that works for ANY org's format.

Flow:
  1. Mechanically extract sheet names, headers, sample rows (fast, no LLM)
  2. Send compact representation to LLM → get structured modelling JSON
  3. Apply modelling to mechanically-parsed form fields
  4. Return SRSData ready for bundle generation
"""

from __future__ import annotations

import json
import logging
import re
from pathlib import Path
from typing import Any

import openpyxl

logger = logging.getLogger(__name__)

# Maximum rows to send per sheet (keeps token count manageable)
_MAX_SAMPLE_ROWS = 8
# Maximum sheets to include in LLM prompt
_MAX_SHEETS = 50


# ---------------------------------------------------------------------------
# Step 1: Mechanical extraction — read sheet structure from XLSX
# ---------------------------------------------------------------------------

def extract_sheet_metadata(file_path: str | Path) -> list[dict[str, Any]]:
    """Extract metadata from every sheet in an XLSX file.

    Returns list of dicts with: name, headers, sample_rows, row_count, col_count.
    No LLM involved — pure cell reading.
    """
    wb = openpyxl.load_workbook(str(file_path), data_only=True)
    sheets: list[dict[str, Any]] = []

    for sheet_name in wb.sheetnames[:_MAX_SHEETS]:
        ws = wb[sheet_name]
        # Read all rows, find the first non-empty row as header
        all_rows: list[list[str]] = []
        for row in ws.iter_rows(
            min_row=1, max_row=min(ws.max_row, _MAX_SAMPLE_ROWS + 5),
            max_col=min(ws.max_column, 20),
            values_only=True,
        ):
            str_row = [str(c).strip() if c is not None else "" for c in row]
            if any(v for v in str_row):
                all_rows.append(str_row)

        if not all_rows:
            continue

        headers = all_rows[0]
        sample_rows = all_rows[1 : _MAX_SAMPLE_ROWS + 1]

        # Count actual data rows (non-empty)
        data_row_count = 0
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, max_col=3, values_only=True):
            if row and any(c is not None for c in row):
                data_row_count += 1

        sheets.append({
            "name": sheet_name,
            "headers": headers,
            "sample_rows": sample_rows,
            "data_rows": data_row_count,
            "col_count": ws.max_column,
        })

    wb.close()
    return sheets


def extract_full_sheet_data(file_path: str | Path, sheet_name: str) -> list[list[str]]:
    """Extract ALL rows from a specific sheet (for form field extraction)."""
    wb = openpyxl.load_workbook(str(file_path), data_only=True)
    ws = wb[sheet_name]
    rows: list[list[str]] = []
    for row in ws.iter_rows(
        min_row=1, max_row=ws.max_row,
        max_col=min(ws.max_column, 20),
        values_only=True,
    ):
        str_row = [str(c).strip() if c is not None else "" for c in row]
        rows.append(str_row)
    wb.close()
    return rows


# ---------------------------------------------------------------------------
# Step 1b: Mechanical modelling extraction — read structured modelling sheets
# ---------------------------------------------------------------------------

# Common header patterns for modelling sheets
_SUBJECT_TYPE_HEADERS = {"subject type name", "subject type", "subjecttype", "name"}
_PROGRAM_HEADERS = {"program name", "program", "name"}
_ENCOUNTER_HEADERS = {"encounter name", "encounter", "name"}

def _find_col(headers: list[str], candidates: set[str]) -> int | None:
    """Find column index matching any candidate (case-insensitive)."""
    for i, h in enumerate(headers):
        if h.strip().lower() in candidates:
            return i
    return None


def extract_modelling_mechanical(file_path: str | Path) -> dict[str, Any] | None:
    """Mechanically extract modelling from a structured modelling XLSX.

    Reads Subject Types, Program, Encounters, Program Encounters sheets
    and returns structured modelling dict. Returns None if no modelling sheets found.
    No LLM involved — pure cell reading.
    """
    wb = openpyxl.load_workbook(str(file_path), data_only=True)
    sheet_names_lower = {s.lower(): s for s in wb.sheetnames}

    modelling: dict[str, Any] = {
        "org_name": "",
        "subject_types": [],
        "programs": [],
        "encounters": [],
        "program_encounters": [],
        "location_hierarchy": [],
    }
    found_any = False

    # --- Subject Types ---
    for key in ("subject types", "subjecttypes", "subject type"):
        if key in sheet_names_lower:
            ws = wb[sheet_names_lower[key]]
            headers = [str(c.value).strip().lower() if c.value else "" for c in list(ws.iter_rows(min_row=1, max_row=1))[0]]
            name_col = _find_col(headers, {"subject type name", "name", "subject type"})
            type_col = _find_col(headers, {"type", "subject type type"})
            if name_col is not None:
                for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
                    vals = [str(c).strip() if c else "" for c in row]
                    name = vals[name_col] if name_col < len(vals) else ""
                    typ = vals[type_col] if type_col is not None and type_col < len(vals) else "Person"
                    if name:
                        modelling["subject_types"].append({"name": name, "type": typ})
                        found_any = True
            break

    # --- Programs ---
    for key in ("program", "programs"):
        if key in sheet_names_lower:
            ws = wb[sheet_names_lower[key]]
            headers = [str(c.value).strip().lower() if c.value else "" for c in list(ws.iter_rows(min_row=1, max_row=1))[0]]
            name_col = _find_col(headers, {"program name", "program", "name"})
            target_col = _find_col(headers, {"target subject type", "subject type", "target subject"})
            enrol_col = _find_col(headers, {"enrolment form", "enrollment form", "enrolment"})
            exit_col = _find_col(headers, {"exit form", "exit"})
            if name_col is not None:
                for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
                    vals = [str(c).strip() if c else "" for c in row]
                    name = vals[name_col] if name_col < len(vals) else ""
                    if name:
                        prog: dict[str, Any] = {"name": name}
                        if target_col is not None and target_col < len(vals) and vals[target_col]:
                            prog["subject_type"] = vals[target_col]
                        if enrol_col is not None and enrol_col < len(vals) and vals[enrol_col]:
                            prog["enrolment_form"] = vals[enrol_col]
                        if exit_col is not None and exit_col < len(vals) and vals[exit_col]:
                            prog["exit_form"] = vals[exit_col]
                        modelling["programs"].append(prog)
                        found_any = True
            break

    # --- Encounters (general, not linked to a program) ---
    for key in ("encounters", "encounter"):
        if key in sheet_names_lower:
            ws = wb[sheet_names_lower[key]]
            headers = [str(c.value).strip().lower() if c.value else "" for c in list(ws.iter_rows(min_row=1, max_row=1))[0]]
            name_col = _find_col(headers, {"encounter name", "encounter", "name"})
            subj_col = _find_col(headers, {"subject type", "subject"})
            freq_col = _find_col(headers, {"frequency", "schedule"})
            sched_col = _find_col(headers, {"encounter type (scheduled/unscheduled)", "encounter type", "scheduled/unscheduled", "type"})
            if name_col is not None:
                for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
                    vals = [str(c).strip() if c else "" for c in row]
                    name = vals[name_col] if name_col < len(vals) else ""
                    if name:
                        enc: dict[str, Any] = {"name": name}
                        if subj_col is not None and subj_col < len(vals) and vals[subj_col]:
                            enc["subject_type"] = vals[subj_col]
                        if freq_col is not None and freq_col < len(vals) and vals[freq_col]:
                            enc["frequency"] = vals[freq_col]
                        if sched_col is not None and sched_col < len(vals) and vals[sched_col]:
                            enc["scheduled"] = vals[sched_col].lower().startswith("sched")
                        modelling["encounters"].append(enc)
                        found_any = True
            break

    # --- Program Encounters ---
    for key in ("program encounters", "programencounters", "program encounter"):
        if key in sheet_names_lower:
            ws = wb[sheet_names_lower[key]]
            headers = [str(c.value).strip().lower() if c.value else "" for c in list(ws.iter_rows(min_row=1, max_row=1))[0]]
            name_col = _find_col(headers, {"encounter name", "encounter", "name"})
            prog_col = _find_col(headers, {"program", "program name"})
            subj_col = _find_col(headers, {"subject type", "subject"})
            if name_col is not None:
                for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
                    vals = [str(c).strip() if c else "" for c in row]
                    name = vals[name_col] if name_col < len(vals) else ""
                    if name:
                        penc: dict[str, Any] = {"name": name}
                        if prog_col is not None and prog_col < len(vals) and vals[prog_col]:
                            penc["program"] = vals[prog_col]
                        if subj_col is not None and subj_col < len(vals) and vals[subj_col]:
                            penc["subject_type"] = vals[subj_col]
                        modelling["program_encounters"].append(penc)
                        found_any = True
            break

    # --- Location Hierarchy ---
    for key in ("location hierarchy", "location hierarchy ", "locations", "address levels"):
        if key in sheet_names_lower:
            ws = wb[sheet_names_lower[key]]
            headers = [str(c.value).strip().lower() if c.value else "" for c in list(ws.iter_rows(min_row=1, max_row=1))[0]]
            name_col = _find_col(headers, {"location type", "address level", "level", "type"})
            if name_col is not None:
                for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
                    vals = [str(c).strip() if c else "" for c in row]
                    name = vals[name_col] if name_col < len(vals) else ""
                    if name:
                        modelling["location_hierarchy"].append(name)
                        found_any = True
            break

    # --- Org name from "Other Important Document" or similar ---
    for key in ("other important document", "overview", "project summary"):
        if key in sheet_names_lower:
            ws = wb[sheet_names_lower[key]]
            for row in ws.iter_rows(min_row=1, max_row=10, max_col=5, values_only=True):
                vals = [str(c).strip() if c else "" for c in row]
                for v in vals:
                    if v and len(v) > 3 and "srs" not in v.lower() and "modelling" not in v.lower() and "form" not in v.lower() and "report" not in v.lower():
                        if not modelling["org_name"]:
                            modelling["org_name"] = v
            break

    wb.close()
    return modelling if found_any else None


def _normalize_subject_type(raw: str, known_types: list[dict[str, str]]) -> str | None:
    """Fuzzy-match a raw subject type string to known subject types."""
    raw_lower = raw.strip().lower()
    for st in known_types:
        name = st.get("name", "").lower()
        if raw_lower == name or raw_lower == name.rstrip("s"):
            return st["name"]
        # Handle typos like "klin" → "Kiln"
        if len(raw_lower) >= 3 and len(name) >= 3:
            # Simple character-level similarity
            common = sum(1 for a, b in zip(sorted(raw_lower), sorted(name)) if a == b)
            if common >= len(raw_lower) * 0.7:
                return st["name"]
    return None


def apply_mechanical_modelling(
    srs_data: "SRSData",
    modelling: dict[str, Any],
) -> "SRSData":
    """Apply mechanically-extracted modelling as ground truth to SRSData.

    This is deterministic — no LLM involved. Uses explicit mappings from
    the modelling document to set subject types, programs, and encounter types.
    """
    from app.models.schemas import SRSData as SRSDataModel

    subject_types = modelling.get("subject_types", [])
    programs = modelling.get("programs", [])
    encounters = modelling.get("encounters", [])
    program_encounters = modelling.get("program_encounters", [])

    # Build lookup tables
    prog_subject: dict[str, str] = {}  # program_name_lower → subject_type_name
    for p in programs:
        st = p.get("subject_type", "")
        if st:
            prog_subject[p["name"].lower()] = st

    enc_subject: dict[str, str] = {}  # encounter_name_lower → subject_type_name
    for e in encounters:
        raw_st = e.get("subject_type", "")
        if raw_st:
            resolved = _normalize_subject_type(raw_st, subject_types)
            if resolved:
                enc_subject[e["name"].lower()] = resolved

    penc_mapping: dict[str, str] = {}  # encounter_name_lower → program_name
    for pe in program_encounters:
        if pe.get("program"):
            penc_mapping[pe["name"].lower()] = pe["program"]

    # Build registration form → subject type mapping
    # "Kiln Reg" or "Kiln Registration" → Kiln
    reg_subject: dict[str, str] = {}
    for st in subject_types:
        st_name = st["name"].lower()
        reg_subject[st_name] = st["name"]

    updated_forms = []
    for form_def in srs_data.forms:
        name_lower = form_def.name.strip().lower()

        # 1. Registration forms: match subject type from form name
        if form_def.formType == "IndividualProfile":
            for st in subject_types:
                st_lower = st["name"].lower()
                if st_lower in name_lower:
                    form_def.subjectTypeName = st["name"]
                    break
            # "Individual Reg" → Individual, "Worker Reg" is NOT registration, it's enrolment

        # 2. Program forms: set subject type from program's target
        if form_def.programName:
            prog_lower = form_def.programName.lower()
            if prog_lower in prog_subject:
                form_def.subjectTypeName = prog_subject[prog_lower]

        # 3. Encounter forms: set subject type from encounters mapping
        if form_def.encounterTypeName:
            enc_lower = form_def.encounterTypeName.lower()
            # Try exact match first, then word-overlap match
            matched_enc_st = enc_subject.get(enc_lower)
            if not matched_enc_st:
                enc_words = set(re.findall(r'\w+', enc_lower))
                for key, st_name in enc_subject.items():
                    key_words = set(re.findall(r'\w+', key))
                    if len(enc_words & key_words) / max(len(enc_words), 1) >= 0.6:
                        matched_enc_st = st_name
                        break
            if matched_enc_st:
                form_def.subjectTypeName = matched_enc_st
            # Check program encounter mapping (exact + fuzzy)
            matched_prog = penc_mapping.get(enc_lower)
            if not matched_prog:
                enc_words = set(re.findall(r'\w+', enc_lower))
                for key, prog_name in penc_mapping.items():
                    key_words = set(re.findall(r'\w+', key))
                    if len(enc_words & key_words) / max(len(enc_words), 1) >= 0.6:
                        matched_prog = prog_name
                        break
            if matched_prog and not form_def.programName:
                form_def.programName = penc_mapping[enc_lower]
                form_def.formType = "ProgramEncounter"
                # Set subject type from program
                prog_lower = form_def.programName.lower()
                if prog_lower in prog_subject:
                    form_def.subjectTypeName = prog_subject[prog_lower]

        # 4. Match general encounter names to encounter sheet
        if form_def.formType == "Encounter" and form_def.encounterTypeName:
            enc_lower = form_def.encounterTypeName.lower()
            enc_words = set(re.findall(r'\w+', enc_lower))
            # Try fuzzy match against encounter names
            for e in encounters:
                e_lower = e["name"].lower()
                e_words = set(re.findall(r'\w+', e_lower))
                # Match if substring match OR significant word overlap
                common_words = enc_words & e_words
                word_overlap = len(common_words) / max(len(enc_words), 1)
                if e_lower in enc_lower or enc_lower in e_lower or word_overlap >= 0.6:
                    raw_st = e.get("subject_type", "")
                    if raw_st:
                        resolved = _normalize_subject_type(raw_st, subject_types)
                        if resolved:
                            form_def.subjectTypeName = resolved
                    # Use full encounter name
                    if len(e["name"]) > len(form_def.encounterTypeName):
                        form_def.encounterTypeName = e["name"]
                    break

        updated_forms.append(form_def)

    # Update the SRSData with modelling data
    return SRSDataModel(
        orgName=modelling.get("org_name") or srs_data.orgName,
        subjectTypes=subject_types if subject_types else srs_data.subjectTypes,
        programs=[p if isinstance(p, dict) else {"name": str(p)} for p in (programs if programs else srs_data.programs)],
        encounterTypes=srs_data.encounterTypes,
        forms=updated_forms,
        groups=srs_data.groups,
        addressLevelTypes=[{"name": lh} for lh in modelling.get("location_hierarchy", [])] or srs_data.addressLevelTypes,
        programEncounterMappings=srs_data.programEncounterMappings,
        generalEncounterTypes=srs_data.generalEncounterTypes,
        visitSchedules=srs_data.visitSchedules,
    )


# ---------------------------------------------------------------------------
# Step 2: Build compact text representation for LLM
# ---------------------------------------------------------------------------

def _build_llm_prompt(sheets: list[dict[str, Any]], file_name: str = "") -> str:
    """Build a MINIMAL text representation of sheets for LLM classification.

    Phase 5 optimization: send ONLY sheet names, headers, and field counts.
    Do NOT send sample row data — it confuses the LLM and causes hallucination.
    The LLM's job is modelling classification (which sheet is what form type),
    not field extraction (that's done mechanically).
    """
    parts: list[str] = []
    parts.append(f"File: {file_name}")
    parts.append(f"Total sheets: {len(sheets)}\n")

    # Detect if any sheet looks like a modelling/summary sheet
    modelling_keywords = {"subject type", "program", "encounter", "form summary", "modelling", "overview"}
    form_keywords = {"field name", "data type", "mandatory", "options", "validation"}

    for i, sheet in enumerate(sheets, 1):
        name = sheet["name"]
        headers_lower = {h.lower() for h in sheet["headers"] if h}
        rows = sheet["data_rows"]

        # Classify sheet type for context
        is_form_sheet = bool(form_keywords & headers_lower)
        is_modelling = bool(modelling_keywords & headers_lower) or any(
            kw in name.lower() for kw in ["summary", "modelling", "model", "program", "subject"]
        )

        sheet_type = "form" if is_form_sheet else ("modelling/metadata" if is_modelling else "unknown")

        parts.append(f"Sheet {i}: \"{name}\" — {rows} rows, type: {sheet_type}")
        parts.append(f"  Headers: {' | '.join(h for h in sheet['headers'] if h)}")

        # For modelling/summary sheets only, include first 3 rows (they contain program/encounter definitions)
        if is_modelling and sheet.get("sample_rows"):
            for j, row in enumerate(sheet["sample_rows"][:3], 1):
                row_str = " | ".join(v for v in row if v)
                if row_str.strip():
                    parts.append(f"  Row {j}: {row_str}")

    return "\n".join(parts)


# ---------------------------------------------------------------------------
# Step 3: LLM classification prompt
# ---------------------------------------------------------------------------

CLASSIFICATION_SYSTEM_PROMPT = """You are an expert at understanding Avni field data collection platform scoping sheets.

Given sheet metadata from an Excel workbook, extract the complete implementation modelling.

Return a JSON object with this EXACT structure:
```json
{
  "org_name": "Organisation name if found, else filename",
  "subject_types": [
    {"name": "Individual", "type": "Person"}
  ],
  "programs": [
    {"name": "Pregnancy", "subject_type": "Individual"},
    {"name": "Child", "subject_type": "Individual"}
  ],
  "forms": [
    {
      "sheet_name": "exact sheet name from input",
      "full_name": "full form name (untruncated)",
      "form_type": "IndividualProfile|ProgramEnrolment|ProgramEncounter|ProgramExit|Encounter|IndividualEncounterCancellation|ProgramEncounterCancellation",
      "subject_type": "subject type this form belongs to",
      "program": "program name or null",
      "encounter_type": "encounter type name or null",
      "in_scope": true
    }
  ],
  "encounter_types": [
    {"name": "ANC", "program": "Pregnancy"},
    {"name": "Delivery", "program": "Pregnancy"}
  ]
}
```

Rules:
1. **form_type classification:**
   - Registration forms → "IndividualProfile" (one per subject type)
   - Enrolment/Enrollment forms → "ProgramEnrolment" (MUST have program name)
   - Exit forms → "ProgramExit" (MUST have program name)
   - Encounter forms WITH a program → "ProgramEncounter" (MUST have program AND encounter_type)
   - Encounter forms WITHOUT a program → "Encounter" (MUST have encounter_type)
   - Cancellation forms for program encounters → "ProgramEncounterCancellation"
   - Cancellation forms for general encounters → "IndividualEncounterCancellation"

2. **CRITICAL — Program linking:**
   - EVERY ProgramEnrolment, ProgramEncounter, and ProgramExit form MUST have a non-null "program" field
   - Infer the program from the form name: "Worker Enrollment" → program: "Worker", "Child Enrollment" → program: "Child", "Pregnancy Enrollment" → program: "Pregnancy"
   - If a form name contains a program keyword (Worker, Child, Pregnancy, etc.), link it to that program
   - Health/Nutrition encounters for children → program: "Child"
   - Health/Nutrition encounters for pregnant women → program: "Pregnancy"

3. **CRITICAL — Subject type assignment:**
   - Each form belongs to exactly ONE subject type
   - Registration forms define which subject type they register (e.g., "Kiln Registration" → subject_type: "Kiln", "Individual Reg" → subject_type: "Individual")
   - Program forms (Enrolment/Encounter/Exit) inherit the subject type from the program's target subject type
   - If a "Subject Types" or "Program" sheet specifies target subject types, use those mappings
   - Worker/Child/Pregnancy programs typically target "Individual" (Person), not "Kiln" or "Family"
   - Encounter forms for a non-Person subject (e.g., "Convergence camp" for Kilns) → subject_type matches that subject

4. **Identifying forms vs metadata sheets:**
   - Form sheets have headers like "Field Name", "Data Type", "Mandatory", "OPTIONS", "Numeric Datatype", "Date Datatype"
   - Metadata sheets: Help, Status Tracker, Project Summary, User persona, W3H, App Dashboard, Visit Scheduling, Reports, Permissions, Review checklist, Location Hierarchy, Subject Types, Program, Encounters, Program Encounters
   - Mark non-form sheets with form_type: "skip"

5. **Scope detection:**
   - Look for "Out of Scope", "Phase 2", "Not yet implemented", "Future" markers
   - Forms after these markers are in_scope: false
   - If a Form Summary/Overview sheet exists, only forms listed there (before any out-of-scope marker) are in_scope: true

6. **Program-encounter mapping:**
   - Use Form Summary, W3H, Modelling, or Program Encounters sheets to determine which encounters belong to which programs
   - W3H sheet typically has form names and scheduling info — use it to understand what forms exist
   - If a sheet name contains a program name, it likely belongs to that program

7. **Sheet name truncation:**
   - Excel limits sheet names to 31 characters
   - If a Form Summary exists, use full names from there
   - Otherwise, infer full name from context (e.g., "ALC (Akshar Learning Center) At" → "ALC (Akshar Learning Center) Attendance")

8. **Subject types:**
   - Default to [{"name": "Individual", "type": "Person"}] unless explicitly different
   - Look for "Subject Type" columns in Form Summary or dedicated Subject Type sheets
   - type values: "Person" (for people), "Individual" (for non-person entities like Kilns, Schools), "Household" (for families/groups), "Group" (for groups)

Return ONLY the JSON object, no markdown fences, no explanation."""


async def classify_sheets_with_llm(
    sheets: list[dict[str, Any]],
    file_name: str = "",
) -> dict[str, Any]:
    """Send sheet metadata to LLM and get structured modelling back.

    Uses the configured LLM provider (Groq/Cerebras for speed).
    """
    from app.services.claude_client import claude_client

    prompt_text = _build_llm_prompt(sheets, file_name)

    user_message = (
        "Analyze this Avni scoping sheet and extract the complete modelling.\n\n"
        f"{prompt_text}\n\n"
        "Return the JSON classification for all sheets."
    )

    response = await claude_client.complete(
        messages=[{"role": "user", "content": user_message}],
        system_prompt=CLASSIFICATION_SYSTEM_PROMPT,
    )

    # Extract JSON from response
    return _parse_llm_json(response)


def _parse_llm_json(text: str) -> dict[str, Any]:
    """Extract JSON object from LLM response (handles markdown fences, etc)."""
    # Try direct parse first
    text = text.strip()
    try:
        return json.loads(text)
    except json.JSONDecodeError:
        pass

    # Try extracting from markdown code fence
    match = re.search(r"```(?:json)?\s*\n?(.*?)\n?```", text, re.DOTALL)
    if match:
        try:
            return json.loads(match.group(1).strip())
        except json.JSONDecodeError:
            pass

    # Try finding first { ... } block
    brace_start = text.find("{")
    if brace_start >= 0:
        depth = 0
        for i in range(brace_start, len(text)):
            if text[i] == "{":
                depth += 1
            elif text[i] == "}":
                depth -= 1
                if depth == 0:
                    try:
                        return json.loads(text[brace_start : i + 1])
                    except json.JSONDecodeError:
                        break

    logger.error("Failed to parse LLM JSON response: %s", text[:500])
    return {}


# ---------------------------------------------------------------------------
# Step 4: Apply LLM modelling to mechanical parser
# ---------------------------------------------------------------------------

def apply_llm_modelling(
    srs_data: "SRSData",
    modelling: dict[str, Any],
) -> "SRSData":
    """Apply LLM-extracted modelling to mechanically-parsed SRSData.

    Updates: form types, program names, encounter type names, and filters scope.
    """
    from app.models.schemas import SRSData as SRSDataModel

    if not modelling:
        logger.warning("No LLM modelling to apply, returning original SRSData")
        return srs_data

    # Build lookup: sheet_name_lower -> modelling entry
    form_lookup: dict[str, dict[str, Any]] = {}
    for f in modelling.get("forms", []):
        sheet_lower = f.get("sheet_name", "").strip().lower()
        if sheet_lower:
            form_lookup[sheet_lower] = f
        # Also index by full_name for matching
        full_lower = f.get("full_name", "").strip().lower()
        if full_lower and full_lower != sheet_lower:
            form_lookup[full_lower] = f

    # Apply to each form
    updated_forms = []
    for form_def in srs_data.forms:
        name_lower = form_def.name.strip().lower()

        # Try exact match, then prefix match
        match = form_lookup.get(name_lower)
        if not match:
            # Try prefix match (truncated names)
            for key, val in form_lookup.items():
                if key.startswith(name_lower[:20]) or name_lower.startswith(key[:20]):
                    match = val
                    break

        if match:
            # Skip out-of-scope forms
            if not match.get("in_scope", True):
                logger.info("Skipping out-of-scope form: %s", form_def.name)
                continue

            # Skip non-form sheets
            if match.get("form_type") == "skip":
                continue

            # Apply full name
            full_name = match.get("full_name", "").strip()
            if full_name and len(full_name) > len(form_def.name.strip()):
                form_def.name = full_name

            # Apply form type
            if match.get("form_type") and match["form_type"] != "skip":
                form_def.formType = match["form_type"]

            # Apply program
            if match.get("program"):
                form_def.programName = match["program"]

            # Apply encounter type
            if match.get("encounter_type"):
                form_def.encounterTypeName = match["encounter_type"]

            # Apply subject type
            if match.get("subject_type"):
                form_def.subjectTypeName = match["subject_type"]

        # Fallback: infer program from form name if formType requires it
        if form_def.formType in ("ProgramEnrolment", "ProgramEncounter", "ProgramExit") and not form_def.programName:
            _name = form_def.name.lower()
            for p in modelling.get("programs", []):
                p_name = p.get("name", "") if isinstance(p, dict) else str(p)
                if p_name.lower() in _name:
                    form_def.programName = p_name
                    break

        updated_forms.append(form_def)

    # Update programs from modelling
    programs = modelling.get("programs", srs_data.programs)
    if isinstance(programs, list) and programs and isinstance(programs[0], dict):
        programs = [p if isinstance(p, dict) else {"name": str(p)} for p in programs]

    # Update subject types
    subject_types = modelling.get("subject_types", srs_data.subjectTypes)

    # Update encounter types from forms
    encounter_types: list[str] = []
    seen: set[str] = set()
    for f in updated_forms:
        if f.encounterTypeName and f.encounterTypeName not in seen:
            encounter_types.append(f.encounterTypeName)
            seen.add(f.encounterTypeName)

    return SRSDataModel(
        orgName=modelling.get("org_name", srs_data.orgName),
        subjectTypes=subject_types,
        programs=programs,
        encounterTypes=encounter_types,
        forms=updated_forms,
        groups=srs_data.groups,
        addressLevelTypes=srs_data.addressLevelTypes,
        programEncounterMappings=srs_data.programEncounterMappings,
        generalEncounterTypes=srs_data.generalEncounterTypes,
        visitSchedules=srs_data.visitSchedules,
    )


# ---------------------------------------------------------------------------
# Main entry point — unified XLSX → SRSData with LLM classification
# ---------------------------------------------------------------------------

async def parse_xlsx_with_llm(file_path: str | Path) -> SRSData:
    """Parse an XLSX file using LLM for modelling + mechanical parser for fields.

    This is the primary entry point. Combines:
    1. Mechanical extraction of all form fields (fast, accurate)
    2. LLM classification of sheets (understands any format)

    Returns SRSData ready for bundle generation.
    """
    from app.services.srs_parser import SRSParser

    file_path = Path(file_path)
    logger.info("LLM-assisted parsing of: %s", file_path.name)

    # Step 1: Extract sheet metadata (fast, no LLM)
    sheets = extract_sheet_metadata(file_path)
    logger.info("Extracted metadata from %d sheets", len(sheets))

    # Step 2: Send to LLM for classification
    modelling = await classify_sheets_with_llm(sheets, file_path.name)
    logger.info(
        "LLM modelling: %d programs, %d encounter types, %d forms",
        len(modelling.get("programs", [])),
        len(modelling.get("encounter_types", [])),
        len(modelling.get("forms", [])),
    )

    # Step 3: Mechanical parse of all form fields
    parser = SRSParser(file_path)
    srs_data = parser.parse()
    logger.info(
        "Mechanical parse: %d forms, %d total fields",
        len(srs_data.forms),
        sum(
            sum(len(g.fields) for g in f.groups)
            for f in srs_data.forms
        ),
    )

    # Step 4: Apply LLM modelling to mechanical parse
    final = apply_llm_modelling(srs_data, modelling)
    logger.info(
        "Final result: %d forms after LLM modelling applied",
        len(final.forms),
    )

    return final
