#!/usr/bin/env python3
"""Ingest SRS scoping documents from Implementations directories into RAG.

Parses Excel SRS files, extracts form definitions, field specifications,
skip logic rules, and implementation patterns, then ingests them as
training data for the RAG pipeline.

Usage:
    python scripts/ingest_srs_docs.py \
        --impl-dirs "/Users/samanvay/Downloads/Implementations" "/Users/samanvay/Downloads/Implementations 2" \
        --output-dir training_data/ \
        --database-url postgresql://avni:avni_ai_dev@localhost:5432/avni_ai

    # Just extract training data (no DB ingestion)
    python scripts/ingest_srs_docs.py \
        --impl-dirs "/Users/samanvay/Downloads/Implementations" "/Users/samanvay/Downloads/Implementations 2" \
        --output-dir training_data/ \
        --no-ingest
"""

import argparse
import glob
import json
import logging
import os
import re
import sys

_SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
_BACKEND_DIR = os.path.dirname(_SCRIPT_DIR)
sys.path.insert(0, _BACKEND_DIR)

logging.basicConfig(level=logging.INFO, format="%(asctime)s %(levelname)s %(message)s")
logger = logging.getLogger("ingest_srs")

# SRS file patterns to look for
SRS_PATTERNS = [
    "*SRS*", "*srs*", "*scoping*", "*Scoping*",
    "*Master*Form*", "*master*form*",
    "*Requirements*App*", "*requirements*app*",
]

# Non-SRS patterns to skip
SKIP_PATTERNS = [
    "report", "Report", "estimate", "Estimate", "budget", "Budget",
    "change log", "Change Log", "tracker", "Tracker", "dashboard", "Dashboard",
    "~$",  # Temp Excel files
]


def find_srs_files(impl_dirs: list[str]) -> list[str]:
    """Find all SRS Excel files in implementation directories."""
    srs_files = []
    for impl_dir in impl_dirs:
        if not os.path.isdir(impl_dir):
            logger.warning("Directory not found: %s", impl_dir)
            continue

        # Walk through all subdirectories
        for root, dirs, files in os.walk(impl_dir):
            for fname in files:
                if not (fname.endswith(".xlsx") or fname.endswith(".xls")):
                    continue

                # Skip temp files and non-SRS patterns
                skip = False
                for pattern in SKIP_PATTERNS:
                    if pattern in fname or pattern in root:
                        skip = True
                        break
                if skip:
                    continue

                # Check if it matches SRS patterns
                full_path = os.path.join(root, fname)
                fname_lower = fname.lower()
                root_lower = root.lower()

                is_srs = False
                for pattern in ["srs", "scoping", "master form", "forms"]:
                    if pattern in fname_lower:
                        is_srs = True
                        break

                # Also check if it's in a "Requirements" or "App" directory
                if not is_srs:
                    if ("requirements" in root_lower and "app" in root_lower) or \
                       ("requirements shared externally" in root_lower):
                        is_srs = True

                if is_srs:
                    srs_files.append(full_path)

    return sorted(set(srs_files))


def parse_srs_excel(filepath: str) -> list[dict]:
    """Parse an SRS Excel file and extract structured data."""
    try:
        import openpyxl
    except ImportError:
        logger.error("openpyxl required: pip install openpyxl")
        return []

    chunks = []
    org_name = _extract_org_name(filepath)

    try:
        wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
    except Exception as e:
        logger.warning("Failed to open %s: %s", filepath, e)
        return chunks

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]

        # Skip non-form sheets
        sheet_lower = sheet_name.lower()
        if any(skip in sheet_lower for skip in [
            "help", "summary", "overview", "status", "report",
            "dashboard", "permission", "privilege", "discussion",
            "user persona", "w3h", "bi report", "card",
        ]):
            continue

        rows = list(ws.iter_rows(values_only=True))
        if len(rows) < 2:
            continue

        # Detect header row
        header_row = None
        header_idx = 0
        for i, row in enumerate(rows[:5]):
            row_strs = [str(c).lower().strip() if c else "" for c in row]
            # Look for common SRS headers
            if any(h in " ".join(row_strs) for h in [
                "field name", "data type", "mandatory", "options",
                "page name", "group", "concept",
            ]):
                header_row = row
                header_idx = i
                break

        if not header_row:
            # Try to extract as free-form text
            text = _sheet_to_text(sheet_name, rows)
            if text and len(text) > 50:
                chunks.append({
                    "content": text,
                    "metadata": {
                        "org": org_name,
                        "sheet": sheet_name,
                        "type": "srs_freeform",
                        "source_file": os.path.basename(filepath),
                    }
                })
            continue

        # Parse as structured SRS
        headers = [str(h).strip().lower() if h else "" for h in header_row]
        col_map = _detect_columns(headers)

        if not col_map.get("field_name"):
            continue

        # Extract fields
        fields = []
        current_group = sheet_name
        for row in rows[header_idx + 1:]:
            if not row or all(c is None or str(c).strip() == "" for c in row):
                continue

            field = {}
            for key, col_idx in col_map.items():
                if col_idx is not None and col_idx < len(row):
                    val = row[col_idx]
                    field[key] = str(val).strip() if val else ""

            # Track group/page changes
            if field.get("group") and field["group"] not in ("", "None"):
                current_group = field["group"]

            field_name = field.get("field_name", "")
            if not field_name or field_name.lower() in ("none", ""):
                continue

            field["group"] = current_group
            fields.append(field)

        if not fields:
            continue

        # Create chunks from the parsed fields
        # Chunk 1: Sheet overview
        field_names = [f.get("field_name", "") for f in fields if f.get("field_name")]
        data_types = set(f.get("data_type", "") for f in fields if f.get("data_type"))
        groups = set(f.get("group", "") for f in fields if f.get("group"))

        overview = (
            f"SRS Form: {sheet_name} (Organization: {org_name})\n"
            f"Total fields: {len(fields)}\n"
            f"Groups/Pages: {', '.join(groups)}\n"
            f"Data types used: {', '.join(data_types)}\n"
            f"Fields: {', '.join(field_names[:30])}"
        )
        if len(field_names) > 30:
            overview += f" and {len(field_names) - 30} more"

        chunks.append({
            "content": overview,
            "metadata": {
                "org": org_name,
                "sheet": sheet_name,
                "type": "srs_overview",
                "field_count": len(fields),
                "source_file": os.path.basename(filepath),
            }
        })

        # Chunk 2: Each field as a training example
        for field in fields:
            field_text = _field_to_text(field, sheet_name, org_name)
            if field_text and len(field_text) > 20:
                chunks.append({
                    "content": field_text,
                    "metadata": {
                        "org": org_name,
                        "sheet": sheet_name,
                        "field_name": field.get("field_name", ""),
                        "data_type": field.get("data_type", ""),
                        "type": "srs_field",
                        "source_file": os.path.basename(filepath),
                    }
                })

        # Chunk 3: Skip logic rules
        for field in fields:
            show_when = field.get("show_when", "")
            hide_when = field.get("hide_when", "")
            if show_when or hide_when:
                rule_text = _rule_to_text(field, sheet_name, org_name)
                if rule_text:
                    chunks.append({
                        "content": rule_text,
                        "metadata": {
                            "org": org_name,
                            "sheet": sheet_name,
                            "field_name": field.get("field_name", ""),
                            "type": "srs_rule",
                            "source_file": os.path.basename(filepath),
                        }
                    })

    try:
        wb.close()
    except Exception:
        pass

    return chunks


def _extract_org_name(filepath: str) -> str:
    """Extract organization name from file path."""
    parts = filepath.split(os.sep)
    for i, part in enumerate(parts):
        if part in ("Implementations", "Implementations 2") and i + 1 < len(parts):
            return parts[i + 1]
    return os.path.basename(os.path.dirname(filepath))


def _detect_columns(headers: list[str]) -> dict:
    """Detect column indices from header names."""
    col_map = {
        "group": None,
        "field_name": None,
        "data_type": None,
        "mandatory": None,
        "options": None,
        "show_when": None,
        "hide_when": None,
        "validation": None,
        "unique": None,
        "default": None,
        "min_val": None,
        "max_val": None,
        "unit": None,
    }

    for i, h in enumerate(headers):
        h = h.strip()
        if not h:
            continue

        if any(x in h for x in ["group", "page name", "page"]) and col_map["group"] is None:
            col_map["group"] = i
        elif any(x in h for x in ["field name", "concept name", "question", "field"]) and "type" not in h:
            if col_map["field_name"] is None:
                col_map["field_name"] = i
        elif any(x in h for x in ["data type", "datatype", "type"]) and "field" not in h:
            if col_map["data_type"] is None:
                col_map["data_type"] = i
        elif any(x in h for x in ["mandatory", "required"]):
            col_map["mandatory"] = i
        elif any(x in h for x in ["option", "answer", "choices"]):
            col_map["options"] = i
        elif any(x in h for x in ["when to show", "show when", "show", "visibility"]):
            if "not" not in h and "hide" not in h:
                col_map["show_when"] = i
        elif any(x in h for x in ["when not to show", "not show", "hide", "hide when"]):
            col_map["hide_when"] = i
        elif "validation" in h:
            col_map["validation"] = i
        elif "unique" in h:
            col_map["unique"] = i
        elif "default" in h:
            col_map["default"] = i
        elif any(x in h for x in ["min", "minimum", "low"]):
            col_map["min_val"] = i
        elif any(x in h for x in ["max", "maximum", "high"]):
            col_map["max_val"] = i
        elif "unit" in h:
            col_map["unit"] = i

    return col_map


def _field_to_text(field: dict, sheet_name: str, org_name: str) -> str:
    """Convert a field dict to a text description for RAG indexing."""
    parts = [f"SRS Field from {org_name} - Form: {sheet_name}"]
    parts.append(f"Field: {field.get('field_name', 'Unknown')}")

    if field.get("data_type"):
        parts.append(f"Data Type: {field['data_type']}")
    if field.get("group"):
        parts.append(f"Group: {field['group']}")
    if field.get("mandatory"):
        parts.append(f"Mandatory: {field['mandatory']}")
    if field.get("options"):
        parts.append(f"Options: {field['options']}")
    if field.get("show_when"):
        parts.append(f"Show when: {field['show_when']}")
    if field.get("hide_when"):
        parts.append(f"Hide when: {field['hide_when']}")
    if field.get("validation"):
        parts.append(f"Validation: {field['validation']}")
    if field.get("min_val"):
        parts.append(f"Min: {field['min_val']}")
    if field.get("max_val"):
        parts.append(f"Max: {field['max_val']}")
    if field.get("unit"):
        parts.append(f"Unit: {field['unit']}")

    return "\n".join(parts)


def _rule_to_text(field: dict, sheet_name: str, org_name: str) -> str:
    """Convert skip logic from SRS into text for RAG indexing."""
    parts = [f"Skip Logic Rule from {org_name} - Form: {sheet_name}"]
    parts.append(f"Target field: {field.get('field_name', 'Unknown')}")

    if field.get("show_when"):
        parts.append(f"SHOW this field when: {field['show_when']}")
    if field.get("hide_when"):
        parts.append(f"HIDE this field when: {field['hide_when']}")

    parts.append(
        "This SRS rule should be converted to an Avni declarative rule or "
        "JavaScript ViewFilter rule using FormElementStatusBuilder."
    )
    return "\n".join(parts)


def _sheet_to_text(sheet_name: str, rows: list) -> str:
    """Convert a non-structured sheet to plain text."""
    lines = []
    for row in rows[:100]:  # Limit to 100 rows
        if row:
            cells = [str(c).strip() for c in row if c is not None and str(c).strip()]
            if cells:
                lines.append(" | ".join(cells))
    return f"Sheet: {sheet_name}\n" + "\n".join(lines)


def generate_srs_training_pairs(chunks: list[dict]) -> list[dict]:
    """Generate additional training pairs from SRS chunks."""
    pairs = []

    for chunk in chunks:
        meta = chunk.get("metadata", {})
        chunk_type = meta.get("type", "")

        if chunk_type == "srs_field":
            field_name = meta.get("field_name", "")
            data_type = meta.get("data_type", "")
            org = meta.get("org", "")
            if field_name and data_type:
                pairs.append({
                    "messages": [
                        {"role": "system", "content": "You are the Avni Platform Architect."},
                        {"role": "user", "content": f"From an SRS for {org}: field '{field_name}' with data type '{data_type}'. What Avni concept type should this be?"},
                        {"role": "assistant", "content": chunk["content"]},
                    ]
                })

        elif chunk_type == "srs_rule":
            field_name = meta.get("field_name", "")
            if field_name:
                pairs.append({
                    "messages": [
                        {"role": "system", "content": "You are the Avni Platform Architect."},
                        {"role": "user", "content": f"Convert this SRS skip logic to Avni format: {chunk['content']}"},
                        {"role": "assistant", "content": (
                            "To convert this SRS skip logic to Avni, you need to create either "
                            "a declarative rule (JSON) or a JavaScript ViewFilter rule. "
                            "The declarative format uses conditions with compoundRule, "
                            "lhs (concept reference with scope), operator, and rhs (answer reference). "
                            "You'll need the exact concept UUIDs from concepts.json."
                        )},
                    ]
                })

    return pairs


def main():
    parser = argparse.ArgumentParser(description="Ingest SRS scoping documents into RAG")
    parser.add_argument("--impl-dirs", nargs="+", required=True,
                        help="Implementation directories to scan")
    parser.add_argument("--output-dir", default="training_data/",
                        help="Output directory for training data")
    parser.add_argument("--database-url", default="",
                        help="PostgreSQL database URL for RAG ingestion")
    parser.add_argument("--no-ingest", action="store_true",
                        help="Skip database ingestion, only generate training data")
    args = parser.parse_args()

    os.makedirs(args.output_dir, exist_ok=True)

    # Step 1: Find SRS files
    logger.info("Scanning for SRS files...")
    srs_files = find_srs_files(args.impl_dirs)
    logger.info("Found %d SRS files", len(srs_files))

    # Step 2: Parse all SRS files
    all_chunks = []
    for filepath in srs_files:
        logger.info("Parsing: %s", os.path.basename(filepath))
        chunks = parse_srs_excel(filepath)
        all_chunks.extend(chunks)
        logger.info("  -> %d chunks", len(chunks))

    logger.info("Total chunks extracted: %d", len(all_chunks))

    # Step 3: Generate training pairs
    training_pairs = generate_srs_training_pairs(all_chunks)
    logger.info("Generated %d training pairs from SRS docs", len(training_pairs))

    # Save training pairs
    srs_train_file = os.path.join(args.output_dir, "avni_srs_train.jsonl")
    with open(srs_train_file, "w") as f:
        for pair in training_pairs:
            f.write(json.dumps(pair) + "\n")
    logger.info("SRS training data saved to: %s", srs_train_file)

    # Save chunks as JSON for RAG ingestion
    chunks_file = os.path.join(args.output_dir, "srs_chunks.json")
    with open(chunks_file, "w") as f:
        json.dump(all_chunks, f, indent=2)
    logger.info("SRS chunks saved to: %s", chunks_file)

    # Step 4: Database ingestion (if requested)
    if not args.no_ingest and args.database_url:
        import asyncio
        asyncio.run(_ingest_to_rag(all_chunks, args.database_url))

    # Summary
    orgs = set(c.get("metadata", {}).get("org", "") for c in all_chunks)
    types = {}
    for c in all_chunks:
        t = c.get("metadata", {}).get("type", "unknown")
        types[t] = types.get(t, 0) + 1

    logger.info("=" * 60)
    logger.info("SRS Ingestion Summary:")
    logger.info("  Files processed: %d", len(srs_files))
    logger.info("  Organizations: %d (%s)", len(orgs), ", ".join(sorted(orgs)[:10]))
    logger.info("  Total chunks: %d", len(all_chunks))
    for t, count in sorted(types.items()):
        logger.info("    %s: %d", t, count)
    logger.info("  Training pairs: %d", len(training_pairs))
    logger.info("=" * 60)


async def _ingest_to_rag(chunks: list[dict], database_url: str):
    """Ingest SRS chunks into pgvector RAG pipeline."""
    from app.services.rag.embeddings import EmbeddingClient
    from app.services.rag.vector_store import VectorStore

    embedding_client = EmbeddingClient()
    vector_store = VectorStore(dsn=database_url)
    await vector_store.initialize()

    # Clear existing SRS data
    await vector_store.clear_collection("srs_examples")

    # Embed and store chunks in batches
    batch_size = 50
    total = 0
    for i in range(0, len(chunks), batch_size):
        batch = chunks[i:i + batch_size]
        texts = [c["content"] for c in batch]
        embeddings = embedding_client.embed_batch(texts)

        db_chunks = []
        for chunk, embedding in zip(batch, embeddings):
            db_chunks.append({
                "collection": "srs_examples",
                "content": chunk["content"],
                "context_prefix": "",
                "embedding": embedding,
                "metadata": chunk.get("metadata", {}),
                "source_file": chunk.get("metadata", {}).get("source_file", ""),
            })

        inserted = await vector_store.upsert_chunks(db_chunks)
        total += inserted
        logger.info("  Ingested batch %d/%d (%d chunks)",
                    (i // batch_size) + 1,
                    (len(chunks) + batch_size - 1) // batch_size,
                    inserted)

    logger.info("RAG ingestion complete: %d chunks in srs_examples collection", total)
    await vector_store.close()


if __name__ == "__main__":
    main()
